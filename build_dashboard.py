#!/usr/bin/env python3
"""Build QuickFlip_Dashboard.html — cyber ops aesthetic."""

import openpyxl, json, os, datetime, re, random, urllib.parse
from collections import Counter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(BASE_DIR, "QuickFlip_Master_Leads.xlsx")
BRIEFS_DIR = os.path.join(BASE_DIR, "Sales Briefs")
OUT_PATH   = os.path.join(BASE_DIR, "Dashboard", "QuickFlip_Dashboard.html")

ZIP_COORDS = {
    "11001":[40.7282,-73.6603],"11005":[40.7193,-73.6999],"11428":[40.7218,-73.7376],
    "11598":[40.6401,-73.7134],"11040":[40.7354,-73.6879],"10038":[40.7074,-74.0021],
    "10003":[40.7297,-73.9890],"10013":[40.7214,-74.0043],"10007":[40.7128,-74.0059],
    "10282":[40.7175,-74.0165],"10014":[40.7358,-74.0042],"10036":[40.7580,-73.9855],
    "10004":[40.7007,-74.0391],"11365":[40.7393,-73.7849],"11010":[40.6962,-73.6784],
    "11530":[40.7218,-73.5973],"11937":[40.9638,-72.1849],"11357":[40.7842,-73.8213],
    "11367":[40.7254,-73.8296],"11375":[40.7217,-73.8480],"11360":[40.7855,-73.7718],
    "11102":[40.7745,-73.9302],"11103":[40.7680,-73.9161],"11354":[40.7674,-73.8305],
    "11358":[40.7565,-73.7869],"11362":[40.7604,-73.7387],"11363":[40.7725,-73.7462],
    "11364":[40.7476,-73.7537],"11366":[40.7232,-73.7723],"11411":[40.6966,-73.7342],
    "11412":[40.6982,-73.7612],"11413":[40.6632,-73.7544],"11418":[40.7015,-73.8308],
    "11420":[40.6762,-73.8151],"11421":[40.6934,-73.8449],"11422":[40.6562,-73.7346],
    "11423":[40.7135,-73.7611],"11426":[40.7441,-73.7242],"11427":[40.7321,-73.7498],
    "11429":[40.7080,-73.7291],"11432":[40.7130,-73.7921],"11434":[40.6691,-73.7791],
    "11435":[40.6957,-73.8026],"11436":[40.6742,-73.7975],"11691":[40.5978,-73.7596],
    "10002":[40.7157,-73.9863],"10009":[40.7264,-73.9796],"10010":[40.7396,-73.9846],
    "10011":[40.7437,-74.0005],"10012":[40.7258,-73.9981],"10016":[40.7477,-73.9821],
    "10017":[40.7530,-73.9760],"10018":[40.7555,-73.9951],"10019":[40.7657,-73.9865],
    "10021":[40.7721,-73.9584],"10022":[40.7601,-73.9665],"10023":[40.7766,-73.9842],
    "10024":[40.7876,-73.9756],"10025":[40.7994,-73.9665],"10026":[40.8026,-73.9549],
    "10027":[40.8117,-73.9560],"10028":[40.7762,-73.9525],"10029":[40.7924,-73.9441],
    "10030":[40.8178,-73.9438],"10031":[40.8239,-73.9486],"10032":[40.8384,-73.9429],
    "10033":[40.8506,-73.9378],"10034":[40.8671,-73.9228],"10040":[40.8588,-73.9318],
    "10065":[40.7632,-73.9634],"10069":[40.7782,-73.9904],"10075":[40.7757,-73.9534],
    "10128":[40.7791,-73.9499],"10280":[40.7077,-74.0158],"10301":[40.6279,-74.0941],
    "11003":[40.6723,-73.7082],"11020":[40.7745,-73.7163],"11021":[40.7890,-73.7279],
    "11023":[40.7829,-73.7357],"11024":[40.8010,-73.6958],"11030":[40.7873,-73.6932],
    "11042":[40.7570,-73.6863],"11050":[40.8248,-73.6847],"11501":[40.7488,-73.6432],
    "11507":[40.7454,-73.6330],"11510":[40.6567,-73.6612],"11514":[40.7238,-73.6119],
    "11516":[40.6301,-73.6978],"11518":[40.6430,-73.6410],"11520":[40.6534,-73.6014],
    "11542":[40.8878,-73.6237],"11545":[40.8348,-73.5965],"11547":[40.8245,-73.5818],
    "11548":[40.8187,-73.5913],"11549":[40.7388,-73.5817],"11557":[40.6497,-73.5999],
    "11559":[40.6223,-73.5950],"11560":[40.8773,-73.5779],"11561":[40.5899,-73.6491],
    "11563":[40.6568,-73.6093],"11566":[40.6621,-73.5527],"11568":[40.7810,-73.5547],
    "11570":[40.6594,-73.6418],"11572":[40.6354,-73.5768],"11576":[40.7989,-73.6432],
    "11577":[40.7922,-73.6533],"11579":[40.8365,-73.5688],"11580":[40.6631,-73.5440],
    "11581":[40.6452,-73.5474],"11590":[40.7549,-73.5832],"11596":[40.7449,-73.5735],
    "94025":[37.4531,-122.1817],"94063":[37.4862,-122.2365],"94303":[37.4513,-122.1180],
    "94105":[37.7879,-122.3963],"94103":[37.7726,-122.4099],"94027":[37.4530,-122.1973],
    "94002":[37.5180,-122.2775],"94010":[37.5688,-122.3699],"94301":[37.4437,-122.1498],
    "94306":[37.4237,-122.1263],"94401":[37.5480,-122.2977],"94402":[37.5399,-122.3199],
    "94403":[37.5310,-122.3016],"94404":[37.5519,-122.2710],"94062":[37.3980,-122.3262],
    "94065":[37.5327,-122.2499],"94066":[37.5236,-122.4029],"94070":[37.5024,-122.2596],
    "33460":[26.6190,-80.0637],"33405":[26.6648,-80.0495],"33401":[26.7153,-80.0534],
    "33407":[26.7445,-80.0671],"33480":[26.7064,-80.0370],"33409":[26.7012,-80.0901],
    "33410":[26.7892,-80.0916],"33411":[26.7091,-80.1369],"33412":[26.7674,-80.1520],
    "33413":[26.6665,-80.1308],"33418":[26.8259,-80.0930],"33426":[26.5899,-80.0820],
    "33435":[26.5280,-80.0712],"33436":[26.5380,-80.0945],"33444":[26.4615,-80.0726],
    "28207":[35.2019,-80.8200],"28205":[35.2177,-80.8052],"28209":[35.1866,-80.8477],
    "28210":[35.1627,-80.8561],"28211":[35.1977,-80.7946],"28270":[35.1176,-80.7749],
    "28277":[35.0612,-80.8208],
    "19035":[40.0447,-75.2710],"19010":[40.0248,-75.3177],"19004":[40.0143,-75.2321],
    "19041":[40.0262,-75.3044],"19066":[40.0193,-75.2521],"19096":[39.9940,-75.2680],
}

def extract_zip(text):
    if not text: return None
    m = re.search(r'\b(\d{5})\b', str(text))
    return m.group(1) if m else None

def get_coords(zip_area):
    z = extract_zip(zip_area)
    if z and z in ZIP_COORDS: return ZIP_COORDS[z]
    t = str(zip_area).lower() if zip_area else ""
    if "manhattan" in t or "nyc" in t: return [40.7580,-73.9855]
    if "queens" in t: return [40.7282,-73.7949]
    if "brooklyn" in t: return [40.6782,-73.9442]
    if "silicon valley" in t or "menlo park" in t or "palo alto" in t: return [37.4274,-122.1697]
    if "redwood city" in t or "peninsula" in t: return [37.4853,-122.2364]
    if "san francisco" in t or "sf" in t: return [37.7749,-122.4194]
    if "palm beach" in t or "west palm" in t or "lake worth" in t: return [26.6614,-80.0711]
    if "charlotte" in t or "myers park" in t: return [35.2271,-80.8431]
    if "gladwyne" in t or "lower merion" in t: return [40.0448,-75.2687]
    if "hamptons" in t or "east hampton" in t or "sagaponack" in t: return [40.9176,-72.2382]
    return None

def safe(v):
    if isinstance(v,(datetime.date,datetime.datetime)): return str(v)
    return "" if v is None else v

def clean(row):
    return {k:safe(v) for k,v in row.items() if k is not None}

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

def sheet_rows(ws, header_row=4):
    headers = [cell.value for cell in ws[header_row]]
    return [dict(zip(headers,row)) for row in ws.iter_rows(min_row=header_row+1,values_only=True) if row[0]]

all_leads    = sheet_rows(wb['Master Leads'])
dev_rows     = sheet_rows(wb['🛠️ Dev Tracker'])
campaign_log = sheet_rows(wb['📊 Campaign Log'])
dev_map      = {str(r['Lead ID']):r for r in dev_rows}

INTEL_DIR = os.path.join(BRIEFS_DIR, 'Intel')
DOCS_DIR  = os.path.join(BASE_DIR, 'Docs')

def parse_sections(md_content):
    """Split markdown into {heading: body} by ## headers."""
    sections = {}
    current = '__intro__'
    buf = []
    for line in md_content.splitlines():
        if line.startswith('## '):
            sections[current] = '\n'.join(buf).strip()
            current = line[3:].strip()
            buf = []
        else:
            buf.append(line)
    sections[current] = '\n'.join(buf).strip()
    return sections

def read_doc(filename):
    path = os.path.join(DOCS_DIR, filename)
    if not os.path.exists(path): return {}
    with open(path, encoding='utf-8') as f: return parse_sections(f.read())

brief_map      = {}   # phone -> filename
brief_sections = {}   # phone -> parsed sections dict
if os.path.exists(BRIEFS_DIR):
    for f in sorted(os.listdir(BRIEFS_DIR)):
        fpath = os.path.join(BRIEFS_DIR, f)
        if f.endswith('.md') and not f.startswith('_') and os.path.isfile(fpath):
            phone_key = f.split('-')[0]
            brief_map[phone_key] = f
            try:
                with open(fpath, encoding='utf-8') as bf:
                    brief_sections[phone_key] = parse_sections(bf.read())
            except Exception:
                brief_sections[phone_key] = {}

intel_map = {}        # phone -> parsed sections dict
if os.path.exists(INTEL_DIR):
    for f in sorted(os.listdir(INTEL_DIR)):
        fpath = os.path.join(INTEL_DIR, f)
        if f.endswith('.md') and os.path.isfile(fpath):
            phone_key = f.split('-')[0]
            try:
                with open(fpath, encoding='utf-8') as inf:
                    intel_map[phone_key] = parse_sections(inf.read())
            except Exception:
                pass

playbook  = read_doc('QF_Sales_Playbook.md')
trade_ref = read_doc('QF_Trade_Value_Reference.md')

for lead in all_leads:
    lid = str(lead.get('Lead ID',''))
    dev = dev_map.get(lid,{})
    lead['Dev Stage']   = dev.get('Stage of Development') or ''
    lead['Staging URL'] = dev.get('Redesign Development Staging Site') or ''
    lead['GitHub Repo'] = dev.get('GitHub Repo Link') or ''
    lead['Template']    = dev.get('Template') or ''
    phone_raw = str(lead.get('Phone','')).replace('(','').replace(')','').replace('-','').replace(' ','').replace('+1','')
    lead['Has Brief']  = phone_raw in brief_map
    lead['Brief File'] = brief_map.get(phone_raw, '')
    lead['Has Intel']  = phone_raw in intel_map
    biz  = str(lead.get('Business Name','') or '')
    addr = str(lead.get('Address','') or lead.get('ZIP / Area','') or '')
    lead['Maps URL'] = 'https://www.google.com/maps/search/?api=1&query=' + urllib.parse.quote_plus((biz+' '+addr).strip())
    coords = get_coords(lead.get('ZIP / Area'))
    if coords:
        random.seed(lid)
        lead['lat'] = round(coords[0]+random.uniform(-0.008,0.008),5)
        lead['lng'] = round(coords[1]+random.uniform(-0.008,0.008),5)
    else:
        lead['lat'] = lead['lng'] = None

# Build campaign ZIP -> income lookup
camp_zip_income = {}
for c in campaign_log:
    z = str(c.get('ZIP Code','') or '').strip()
    if z: camp_zip_income[z] = str(c.get('Per Capita Income','') or '')

def norm_url(raw):
    raw = str(raw or '').strip()
    if not raw or raw.lower() in ('no','n/a','—','-','none',''): return ''
    if raw.startswith('http'): return raw
    if '.' in raw and ' ' not in raw: return 'https://' + raw.lstrip('/')
    return ''

# Build SALES_PORTALS: one entry per lead that has a brief
sales_portals = {}
for lead in all_leads:
    phone_raw = str(lead.get('Phone','')).replace('(','').replace(')','').replace('-','').replace(' ','').replace('+1','')
    if phone_raw not in brief_map:
        continue
    lid      = str(lead.get('Lead ID',''))
    zip_area = str(lead.get('ZIP / Area','') or '')
    zip_income = next((inc for z,inc in camp_zip_income.items() if z in zip_area), '')
    website_raw = str(lead.get('Website','') or '')
    brief_intro = brief_sections.get(phone_raw, {}).get('__intro__', '')
    brief_meta = {mm.group(1).strip(): mm.group(2).strip() for mm in re.finditer(r'\*\*([^*:]+):\*\*\s*(.*)', brief_intro)}
    sales_portals[lid] = {
        'biz':       str(lead.get('Business Name','') or ''),
        'trade':     str(lead.get('Trade','') or ''),
        'phone':     str(lead.get('Phone','') or ''),
        'address':   str(lead.get('Address','') or ''),
        'zip_area':  zip_area,
        'rating':    lead.get('Rating ★',''),
        'reviews':   lead.get('Reviews',''),
        'website':   norm_url(website_raw),
        'demo_url':  str(lead.get('Staging URL','') or ''),
        'maps_url':  str(lead.get('Maps URL','') or ''),
        'dev_stage': str(lead.get('Dev Stage','') or ''),
        'zip_income':zip_income,
        'brief':     brief_sections.get(phone_raw, {}),
        'intel':     intel_map.get(phone_raw, None),
        'has_intel': phone_raw in intel_map,
        'owner':     brief_meta.get('Owner', ''),
        'email':     brief_meta.get('Email', ''),
        'website_brief': norm_url(brief_meta.get('Website', '')),
    }

leads_clean    = [clean(r) for r in all_leads]
dev_clean      = [clean(r) for r in dev_rows]
campaign_clean = [clean(r) for r in campaign_log]

total       = len(all_leads)
yes_count   = sum(1 for l in all_leads if l['🎯 Target?']=='✅ Yes')
maybe_count = sum(1 for l in all_leads if l['🎯 Target?']=='🤔 Maybe')
no_count    = sum(1 for l in all_leads if l['🎯 Target?']=='❌ No')
pitch_ready = sum(1 for l in all_leads if l['Dev Stage']=='✅ Completed')
in_dev      = sum(1 for l in all_leads if l['Dev Stage'] in ['🔨 Base Redesign to Template','🎨 Colors & Logos Updated'])
pending_dev = sum(1 for l in all_leads if l['🎯 Target?']=='✅ Yes' and l['Dev Stage'] not in ['✅ Completed','🔨 Base Redesign to Template','🎨 Colors & Logos Updated'])

trade_counts_yes = dict(Counter(l['Trade'] for l in all_leads if l['🎯 Target?']=='✅ Yes' and l.get('Trade')))

kpis = {"total":total,"yes":yes_count,"maybe":maybe_count,"no":no_count,
        "pitch_ready":pitch_ready,"in_dev":in_dev,"pending_dev":pending_dev,
        "clients_won":0,"trade_counts_yes":trade_counts_yes}

def jdump(obj):
    """JSON-encode and escape </ so embedded markdown can't break the script tag."""
    return json.dumps(obj, ensure_ascii=False).replace('</', '<\\/')

data_js = f"""const LEADS={jdump(leads_clean)};
const DEV_ROWS={jdump(dev_clean)};
const CAMPAIGN_LOG={jdump(campaign_clean)};
const KPIS={jdump(kpis)};
const SALES_PORTALS={jdump(sales_portals)};
const PLAYBOOK={jdump(playbook)};
const TRADE_REF={jdump(trade_ref)};
const REFRESHED="{datetime.date.today()}";
"""


HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>QuickFlip Sites — Ops Dashboard</title>
<link rel="preconnect" href="https://fonts.googleapis.com"/>
<link href="https://fonts.googleapis.com/css2?family=Rajdhani:wght@400;500;600;700&family=Share+Tech+Mono&display=swap" rel="stylesheet"/>
<script src="https://cdn.tailwindcss.com"></script>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<script src="https://unpkg.com/leaflet.heat@0.2.0/dist/leaflet-heat.js"></script>
<script src="https://cdn.jsdelivr.net/npm/marked@9/marked.min.js"></script>
<style>
:root{
  --c:    #00d4ff;
  --c2:   #0090cc;
  --g:    #00ff9d;
  --p:    #7b5ea7;
  --o:    #ff6b35;
  --y:    #f5c400;
  --bg:   #060b18;
  --bg2:  #0a1020;
  --panel:#0d1528;
  --border:rgba(0,212,255,.22);
  --glow: rgba(0,212,255,.12);
}
*{box-sizing:border-box;margin:0;padding:0}
body{
  font-family:'Rajdhani',sans-serif;
  background:var(--bg);
  color:#c8d8e8;
  min-height:100vh;
  overflow-x:hidden;
}

/* animated grid bg */
body::before{
  content:'';position:fixed;inset:0;
  background-image:
    linear-gradient(rgba(0,212,255,.03) 1px,transparent 1px),
    linear-gradient(90deg,rgba(0,212,255,.03) 1px,transparent 1px);
  background-size:40px 40px;
  pointer-events:none;z-index:0;
}

/* ── scan line ── */
@keyframes scan{0%{top:-4px}100%{top:100%}}
body::after{
  content:'';position:fixed;left:0;width:100%;height:3px;
  background:linear-gradient(transparent,rgba(0,212,255,.18),transparent);
  animation:scan 6s linear infinite;pointer-events:none;z-index:1;
}

/* ── Header ── */
#hdr{
  position:relative;z-index:10;
  background:linear-gradient(180deg,rgba(0,20,50,.98),rgba(6,11,24,.95));
  border-bottom:1px solid var(--border);
  padding:10px 24px;
  display:flex;align-items:center;justify-content:space-between;
  box-shadow:0 0 30px rgba(0,212,255,.08);
}
#hdr h1{
  font-size:20px;font-weight:700;letter-spacing:3px;text-transform:uppercase;
  color:var(--c);text-shadow:0 0 12px rgba(0,212,255,.7);
}
#hdr .sub{font-size:11px;color:#4a6080;letter-spacing:2px;margin-top:2px}
.ts{font-family:'Share Tech Mono',monospace;font-size:12px;color:#3a7090;letter-spacing:1px}
.ts span{color:var(--c)}

/* ── Tab nav ── */
#tabs{
  position:relative;z-index:10;
  display:flex;gap:0;
  background:rgba(6,11,24,.95);
  border-bottom:1px solid var(--border);
  padding:0 16px;
  overflow-x:auto;
}
.tab-btn{
  padding:11px 20px;font-size:13px;font-weight:600;
  letter-spacing:1px;text-transform:uppercase;
  color:#3a5570;border:none;background:transparent;cursor:pointer;
  border-bottom:2px solid transparent;
  transition:all .2s;white-space:nowrap;
}
.tab-btn:hover{color:#0090cc}
.tab-btn.active{color:var(--c);border-bottom-color:var(--c);text-shadow:0 0 8px rgba(0,212,255,.5)}

/* ── Panel card ── */
.panel{
  position:relative;
  background:var(--panel);
  border:1px solid var(--border);
  border-radius:4px;
  box-shadow:0 0 20px var(--glow),inset 0 0 40px rgba(0,212,255,.02);
  overflow:hidden;
}
/* corner accents */
.panel::before,.panel::after{
  content:'';position:absolute;width:12px;height:12px;
  border-color:var(--c);border-style:solid;opacity:.7;
}
.panel::before{top:0;left:0;border-width:2px 0 0 2px}
.panel::after{bottom:0;right:0;border-width:0 2px 2px 0}

.panel-title{
  font-size:11px;font-weight:600;letter-spacing:2px;text-transform:uppercase;
  color:#3a7090;padding:10px 14px 6px;
  border-bottom:1px solid rgba(0,212,255,.08);
}
.panel-title span{color:var(--c)}

/* ── KPI cards ── */
.kpi{
  padding:14px 16px;
  background:var(--panel);
  border:1px solid var(--border);
  border-radius:4px;
  box-shadow:0 0 16px var(--glow);
  position:relative;overflow:hidden;
}
.kpi::before{
  content:'';position:absolute;top:0;left:0;right:0;height:1px;
  background:linear-gradient(90deg,transparent,var(--c),transparent);
  opacity:.6;
}
.kpi-val{
  font-family:'Share Tech Mono',monospace;
  font-size:44px;font-weight:700;line-height:1;
  color:var(--c);text-shadow:0 0 14px rgba(0,212,255,.6);
}
.kpi-val.green{color:var(--g);text-shadow:0 0 14px rgba(0,255,157,.5)}
.kpi-val.yellow{color:var(--y);text-shadow:0 0 14px rgba(245,196,0,.5)}
.kpi-val.orange{color:var(--o);text-shadow:0 0 14px rgba(255,107,53,.5)}
.kpi-val.purple{color:#a78bfa;text-shadow:0 0 14px rgba(167,139,250,.5)}
.kpi-val.red{color:#f87171;text-shadow:0 0 14px rgba(248,113,113,.4)}
.kpi-lbl{font-size:10px;letter-spacing:2px;text-transform:uppercase;color:#3a5570;margin-top:4px}

/* ── Tables ── */
.tbl-wrap{overflow:auto;max-height:220px}
.tbl-wrap::-webkit-scrollbar{width:4px;height:4px}
.tbl-wrap::-webkit-scrollbar-track{background:transparent}
.tbl-wrap::-webkit-scrollbar-thumb{background:rgba(0,212,255,.2);border-radius:2px}
table{width:100%;border-collapse:collapse;font-size:12px}
th{
  background:rgba(0,212,255,.05);
  padding:7px 10px;text-align:left;
  font-size:10px;font-weight:600;letter-spacing:1.5px;text-transform:uppercase;
  color:#2a6080;position:sticky;top:0;z-index:2;
  border-bottom:1px solid rgba(0,212,255,.1);
  cursor:pointer;white-space:nowrap;
}
th:hover{color:var(--c)}
td{padding:6px 10px;border-bottom:1px solid rgba(0,212,255,.05);vertical-align:middle}
tr:hover td{background:rgba(0,212,255,.04)}

/* ── Badges ── */
.badge{display:inline-block;padding:2px 7px;border-radius:2px;font-size:10px;font-weight:600;letter-spacing:.5px}
.b-done{background:rgba(0,255,157,.12);color:var(--g);border:1px solid rgba(0,255,157,.3)}
.b-base{background:rgba(245,196,0,.1);color:var(--y);border:1px solid rgba(245,196,0,.3)}
.b-colors{background:rgba(167,139,250,.1);color:#a78bfa;border:1px solid rgba(167,139,250,.3)}
.b-pending{background:rgba(58,80,112,.2);color:#3a5070;border:1px solid rgba(58,80,112,.3)}
.b-nosite{background:rgba(255,107,53,.1);color:var(--o);border:1px solid rgba(255,107,53,.3)}
.b-yes{background:rgba(0,255,157,.08);color:var(--g);border:1px solid rgba(0,255,157,.2)}
.b-maybe{background:rgba(0,144,204,.1);color:var(--c2);border:1px solid rgba(0,144,204,.3)}

/* ── Controls ── */
.ctrl{
  background:rgba(0,212,255,.05);border:1px solid rgba(0,212,255,.2);
  border-radius:3px;padding:5px 10px;color:#8ab0c8;font-size:12px;
  font-family:'Rajdhani',sans-serif;letter-spacing:.5px;
}
.ctrl:focus{outline:none;border-color:var(--c);box-shadow:0 0 8px rgba(0,212,255,.2)}
.tog{
  background:rgba(0,212,255,.05);border:1px solid rgba(0,212,255,.2);
  border-radius:3px;padding:5px 12px;font-size:11px;letter-spacing:1px;
  text-transform:uppercase;cursor:pointer;color:#3a6080;
  font-family:'Rajdhani',sans-serif;transition:all .15s;
}
.tog.on{background:rgba(0,212,255,.12);border-color:var(--c);color:var(--c);box-shadow:0 0 8px rgba(0,212,255,.2)}

/* ── Map ── */
#ov-map,#full-map{border-radius:4px;border:1px solid var(--border)}
#ov-map{height:290px}
#full-map{height:520px}
.leaflet-container{background:#060b18 !important}
.leaflet-popup-content-wrapper{
  background:rgba(10,16,32,.97);color:#c8d8e8;
  border:1px solid rgba(0,212,255,.3);border-radius:4px;
  box-shadow:0 0 20px rgba(0,212,255,.2);
}
.leaflet-popup-tip{background:rgba(10,16,32,.97)}
.leaflet-control-zoom a{
  background:rgba(6,11,24,.9)!important;color:var(--c)!important;
  border-color:rgba(0,212,255,.3)!important;
}

/* ── Revenue pill ── */
.rev-pill{
  background:linear-gradient(90deg,rgba(0,255,157,.08),rgba(0,212,255,.06));
  border:1px solid rgba(0,255,157,.25);border-radius:4px;
  padding:10px 16px;font-size:13px;font-weight:600;letter-spacing:1px;
  color:var(--g);text-shadow:0 0 8px rgba(0,255,157,.4);
}

/* ── Ranked list (overview sidebar) ── */
.rank-item{
  display:flex;align-items:center;gap:10px;padding:7px 12px;
  border-bottom:1px solid rgba(0,212,255,.05);
  transition:background .15s;
}
.rank-item:hover{background:rgba(0,212,255,.04)}
.rank-num{
  font-family:'Share Tech Mono',monospace;font-size:11px;
  color:rgba(0,212,255,.35);width:20px;text-align:right;flex-shrink:0;
}
.rank-bar-wrap{flex:1;height:4px;background:rgba(0,212,255,.08);border-radius:2px;overflow:hidden}
.rank-bar{height:100%;background:linear-gradient(90deg,var(--c2),var(--c));border-radius:2px}
.rank-val{
  font-family:'Share Tech Mono',monospace;font-size:11px;
  color:var(--c);width:24px;text-align:right;flex-shrink:0;
}

/* ── Overview grid ── */
#ov-grid{
  display:grid;
  grid-template-columns:230px 1fr 250px;
  grid-template-rows:auto auto;
  gap:8px;
  padding:10px 14px 14px;
  position:relative;z-index:2;
}
.ov-kpi-row{grid-column:1/-1;display:grid;grid-template-columns:repeat(7,1fr);gap:8px}
.ov-left{grid-column:1;grid-row:2;display:flex;flex-direction:column;gap:8px}
.ov-center{grid-column:2;grid-row:2;display:flex;flex-direction:column;gap:8px}
.ov-right{grid-column:3;grid-row:2;display:flex;flex-direction:column;gap:8px}
.ov-map-panel{display:flex;flex-direction:column}
.ov-map-panel .panel-title{flex-shrink:0}
.ov-map-inner{padding:8px}

/* Tab content areas */
.tab-content{display:none;position:relative;z-index:2}
.tab-content.active{display:block}
.tab-pad{padding:14px;display:flex;flex-direction:column;gap:10px}

/* controls row */
.ctrl-row{display:flex;gap:8px;align-items:center;flex-wrap:wrap}

/* pulsing dot */
@keyframes pulse{0%,100%{opacity:1;transform:scale(1)}50%{opacity:.5;transform:scale(.8)}}
.dot-live{width:7px;height:7px;border-radius:50%;background:var(--g);
  animation:pulse 2s infinite;box-shadow:0 0 6px var(--g);display:inline-block}
</style>
</head>
<body>

<!-- ── HEADER ── -->
<div id="hdr">
  <div>
    <h1>⚡ QuickFlip Sites</h1>
    <div class="sub">OPERATIONS COMMAND CENTER</div>
  </div>
  <div style="display:flex;align-items:center;gap:16px">
    <span class="dot-live"></span>
    <div class="ts">REFRESHED <span id="refreshed"></span></div>
  </div>
</div>

<!-- ── TABS ── -->
<div id="tabs">
  <button class="tab-btn active" data-tab="overview">◈ Overview</button>
  <button class="tab-btn" data-tab="dev">⬡ Dev Tracker</button>
  <button class="tab-btn" data-tab="sales">▶ Sales Queue</button>
  <button class="tab-btn" data-tab="pipeline">↻ Active Pipeline</button>
  <button class="tab-btn" data-tab="maybe">◎ Maybe</button>
  <button class="tab-btn" data-tab="campaign">▦ Campaigns</button>
  <button class="tab-btn" data-tab="map">⊕ Full Map</button>
</div>

<!-- ══════════════ TAB: OVERVIEW ══════════════ -->
<div id="tab-overview" class="tab-content active">
<div id="ov-grid">

  <!-- KPI row -->
  <div class="ov-kpi-row">
    <div class="kpi"><div class="kpi-val" id="kv-total"></div><div class="kpi-lbl">Total Leads</div></div>
    <div class="kpi"><div class="kpi-val" style="color:#c084fc;text-shadow:0 0 14px rgba(192,132,252,.55)" id="kv-yes"></div><div class="kpi-lbl">Yes Targets</div></div>
    <div class="kpi"><div class="kpi-val" style="color:#00ff9d;text-shadow:0 0 14px rgba(0,255,157,.5)" id="kv-pitch"></div><div class="kpi-lbl">Pitch Ready</div></div>
    <div class="kpi"><div class="kpi-val yellow" id="kv-indev"></div><div class="kpi-lbl">In Dev</div></div>
    <div class="kpi"><div class="kpi-val" style="color:#4a6080" id="kv-pend"></div><div class="kpi-lbl">Pending Dev</div></div>
    <div class="kpi"><div class="kpi-val" style="color:#4fa3cf" id="kv-maybe"></div><div class="kpi-lbl">Maybe</div></div>
    <div class="kpi"><div class="kpi-val red" id="kv-clients"></div><div class="kpi-lbl">Clients Won</div></div>
  </div>

  <!-- LEFT: insights on top, sales queue below -->
  <div class="ov-left">
    <div class="panel" style="overflow-y:auto;max-height:420px">
      <div class="panel-title" style="position:sticky;top:0;background:var(--panel);z-index:1">◈ <span>Data Insights</span></div>
      <div id="ov-insights"></div>
    </div>
    <div class="panel">
      <div class="panel-title">▶ <span>Sales Queue</span> — Pitch Ready</div>
      <div id="ov-sales-list"></div>
    </div>
  </div>

  <!-- CENTER: Map + campaign bar below -->
  <div class="ov-center">
    <div class="panel ov-map-panel">
      <div class="panel-title">⊕ <span>Lead Coverage Map</span> — All Targets</div>
      <div class="ov-map-inner"><div id="ov-map"></div></div>
    </div>
    <div class="panel">
      <div class="panel-title">▶ <span>Sales Queue</span> — Quick Access</div>
      <div style="overflow-x:auto">
        <table id="ov-sq-table" style="font-size:11px">
          <thead><tr>
            <th>Business</th><th>Trade</th><th>Phone</th><th>Preview</th><th>Brief</th><th>Maps</th>
          </tr></thead>
          <tbody id="ov-sq-body"></tbody>
        </table>
      </div>
    </div>
  </div>

  <!-- RIGHT: pipeline funnel + trade breakdown + stage bars -->
  <div class="ov-right">
    <div class="panel" style="padding:10px 12px">
      <div class="panel-title" style="padding:0 0 8px">▸ <span>Pipeline Funnel</span></div>
      <canvas id="ov-funnel" height="110"></canvas>
    </div>
    <div class="panel" style="padding:10px 12px">
      <div class="panel-title" style="padding:0 0 8px">▸ <span>Yes Leads by Trade</span></div>
      <canvas id="ov-trade" height="130"></canvas>
    </div>
    <div class="panel">
      <div class="panel-title">▦ <span>Dev Stage</span> Breakdown</div>
      <div id="ov-stage-bars" style="padding:10px 14px"></div>
    </div>
  </div>

</div><!-- /ov-grid -->
</div>

<!-- ══════════════ TAB: DEV TRACKER ══════════════ -->
<div id="tab-dev" class="tab-content">
<div class="tab-pad">
  <div class="ctrl-row">
    <input class="ctrl" type="text" id="devSearch" placeholder="Search name / trade / ZIP…" oninput="filterDev()" style="width:230px"/>
    <select class="ctrl" id="devStage" onchange="filterDev()">
      <option value="">All Stages</option>
      <option>⏳ Pending</option>
      <option>🔨 Base Redesign to Template</option>
      <option>🎨 Colors &amp; Logos Updated</option>
      <option>✅ Completed</option>
    </select>
    <select class="ctrl" id="devTrade" onchange="filterDev()"><option value="">All Trades</option></select>
    <button class="tog" id="devNoSite" onclick="toggleBtn('devNoSite');filterDev()">No Website Only</button>
    <span style="margin-left:auto;font-size:11px;letter-spacing:1px;color:#2a5070" id="devCount"></span>
  </div>
  <div class="panel"><div class="tbl-wrap" style="max-height:calc(100vh - 220px)">
    <table id="devTable"><thead><tr>
      <th onclick="sortTable('devTable',0)">ID</th>
      <th onclick="sortTable('devTable',1)">Business Name</th>
      <th onclick="sortTable('devTable',2)">Trade</th>
      <th onclick="sortTable('devTable',3)">ZIP / Area</th>
      <th onclick="sortTable('devTable',4)">Phone</th>
      <th onclick="sortTable('devTable',5)">Stage</th>
      <th>Staging URL</th>
      <th onclick="sortTable('devTable',7)">Template</th>
      <th>Current Site</th>
    </tr></thead><tbody></tbody></table>
  </div></div>
</div>
</div>

<!-- ══════════════ TAB: SALES QUEUE ══════════════ -->
<div id="tab-sales" class="tab-content">
<div class="tab-pad">
  <div class="rev-pill" id="revenuePill"></div>
  <div class="ctrl-row">
    <input class="ctrl" type="text" id="salesSearch" placeholder="Search…" oninput="filterSales()" style="width:200px"/>
    <select class="ctrl" id="salesTrade" onchange="filterSales()"><option value="">All Trades</option></select>
    <span style="margin-left:auto;font-size:11px;letter-spacing:1px;color:#2a5070" id="salesCount"></span>
  </div>
  <div class="panel"><div class="tbl-wrap" style="max-height:calc(100vh - 240px)">
    <table id="salesTable"><thead><tr>
      <th onclick="sortTable('salesTable',0)">Business Name</th>
      <th onclick="sortTable('salesTable',1)">Trade</th>
      <th onclick="sortTable('salesTable',2)">ZIP / Area</th>
      <th onclick="sortTable('salesTable',3)">Phone</th>
      <th>Staging Site</th>
      <th>Portal / Intel</th>
      <th onclick="sortTable('salesTable',6)">Rating</th>
      <th onclick="sortTable('salesTable',7)">Reviews</th>
      <th>Status</th>
    </tr></thead><tbody></tbody></table>
  </div></div>
</div>
</div>

<!-- ══════════════ TAB: ACTIVE PIPELINE ══════════════ -->
<div id="tab-pipeline" class="tab-content">
<div class="tab-pad">
  <div class="ctrl-row">
    <input class="ctrl" type="text" id="pipeSearch" placeholder="Search…" oninput="filterPipe()" style="width:200px"/>
    <select class="ctrl" id="pipeStage" onchange="filterPipe()">
      <option value="">All Active Stages</option>
      <option>🔨 Base Redesign to Template</option>
      <option>🎨 Colors &amp; Logos Updated</option>
    </select>
    <select class="ctrl" id="pipeTrade" onchange="filterPipe()"><option value="">All Trades</option></select>
    <span style="margin-left:auto;font-size:11px;letter-spacing:1px;color:#2a5070" id="pipeCount"></span>
  </div>
  <div class="panel"><div class="tbl-wrap" style="max-height:calc(100vh - 220px)">
    <table id="pipeTable"><thead><tr>
      <th onclick="sortTable('pipeTable',0)">Business Name</th>
      <th onclick="sortTable('pipeTable',1)">Trade</th>
      <th onclick="sortTable('pipeTable',2)">ZIP / Area</th>
      <th onclick="sortTable('pipeTable',3)">Phone</th>
      <th onclick="sortTable('pipeTable',4)">Stage</th>
      <th onclick="sortTable('pipeTable',5)">Template</th>
      <th>Staging URL</th>
      <th>GitHub</th>
      <th>Current Site</th>
    </tr></thead><tbody></tbody></table>
  </div></div>
</div>
</div>

<!-- ══════════════ TAB: MAYBE ══════════════ -->
<div id="tab-maybe" class="tab-content">
<div class="tab-pad">
  <div class="ctrl-row">
    <input class="ctrl" type="text" id="maybeSearch" placeholder="Search…" oninput="filterMaybe()" style="width:200px"/>
    <select class="ctrl" id="maybeTrade" onchange="filterMaybe()"><option value="">All Trades</option></select>
    <button class="tog" id="maybeNoSite" onclick="toggleBtn('maybeNoSite');filterMaybe()">No Website Only</button>
    <span style="margin-left:auto;font-size:11px;letter-spacing:1px;color:#2a5070" id="maybeCount"></span>
  </div>
  <div class="panel"><div class="tbl-wrap" style="max-height:calc(100vh - 220px)">
    <table id="maybeTable"><thead><tr>
      <th onclick="sortTable('maybeTable',0)">Business Name</th>
      <th onclick="sortTable('maybeTable',1)">Trade</th>
      <th onclick="sortTable('maybeTable',2)">ZIP / Area</th>
      <th onclick="sortTable('maybeTable',3)">Rating</th>
      <th onclick="sortTable('maybeTable',4)">Reviews</th>
      <th onclick="sortTable('maybeTable',5)">Has Website?</th>
      <th>Notes / Intel</th>
      <th>Maps</th>
    </tr></thead><tbody></tbody></table>
  </div></div>
</div>
</div>

<!-- ══════════════ TAB: CAMPAIGNS ══════════════ -->
<div id="tab-campaign" class="tab-content">
<div class="tab-pad">
  <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px">
    <div class="panel" style="padding:14px">
      <div class="panel-title" style="padding:0 0 10px">▦ <span>Qualified Leads</span> per ZIP Campaign</div>
      <canvas id="campChart" height="200"></canvas>
    </div>
    <div class="panel" style="padding:14px">
      <div class="panel-title" style="padding:0 0 10px">▦ <span>Per Capita Income</span> by ZIP</div>
      <canvas id="incomeChart" height="200"></canvas>
    </div>
  </div>
  <div class="panel"><div class="tbl-wrap" style="max-height:calc(100vh - 400px)">
    <table id="campTable"><thead><tr>
      <th onclick="sortTable('campTable',0)">ZIP</th>
      <th onclick="sortTable('campTable',1)">Location</th>
      <th onclick="sortTable('campTable',2)">State</th>
      <th onclick="sortTable('campTable',3)">Per Capita Income</th>
      <th onclick="sortTable('campTable',4)">Date Searched</th>
      <th>Trades Searched</th>
      <th onclick="sortTable('campTable',6)">Leads Found</th>
      <th onclick="sortTable('campTable',7)">Qualified</th>
    </tr></thead><tbody></tbody></table>
  </div></div>
</div>
</div>

<!-- ══════════════ TAB: FULL MAP ══════════════ -->
<div id="tab-map" class="tab-content">
<div class="tab-pad">
  <div class="ctrl-row">
    <select class="ctrl" id="mapTarget" onchange="renderFullMap()">
      <option value="all">All Targets</option>
      <option value="yes">Yes Only</option>
      <option value="maybe">Maybe Only</option>
      <option value="pitch">Pitch Ready Only</option>
    </select>
    <select class="ctrl" id="mapTrade" onchange="renderFullMap()"><option value="">All Trades</option></select>
    <button class="tog on" id="mapShowNo" onclick="toggleBtn('mapShowNo');renderFullMap()">Show No/Pass</button>
    <button class="tog on" id="mapHeat" onclick="toggleBtn('mapHeat');toggleHeat()">Heat Overlay</button>
    <div style="display:flex;gap:12px;margin-left:12px;font-size:11px;letter-spacing:.5px;color:#3a5570">
      <span style="color:#00ff9d">● Pitch Ready</span>
      <span style="color:#f5c400">● In Dev</span>
      <span style="color:#4f46e5">● Pending</span>
      <span style="color:#a78bfa">● Maybe</span>
      <span style="color:#3a5070">● No/Pass</span>
    </div>
  </div>
  <div class="panel" style="padding:8px">
    <div id="full-map"></div>
  </div>
</div>
</div>

<!-- ══════════════ SCRIPTS ══════════════ -->
<script>
__DATA_PLACEHOLDER__
</script>
<script>
// ── Tabs ─────────────────────────────────────────────────────────────────────
document.querySelectorAll('.tab-btn').forEach(btn=>{
  btn.addEventListener('click',()=>{
    document.querySelectorAll('.tab-btn').forEach(b=>b.classList.remove('active'));
    document.querySelectorAll('.tab-content').forEach(c=>c.classList.remove('active'));
    btn.classList.add('active');
    document.getElementById('tab-'+btn.dataset.tab).classList.add('active');
    if(btn.dataset.tab==='map') setTimeout(initFullMap,100);
    if(btn.dataset.tab==='campaign') setTimeout(buildCampaignCharts,100);
  });
});

// ── Helpers ───────────────────────────────────────────────────────────────────
function toggleBtn(id){document.getElementById(id).classList.toggle('on')}
function isOn(id){return document.getElementById(id).classList.contains('on')}

function stageBadge(s){
  if(!s) return '<span class="badge b-pending">⏳ Pending</span>';
  if(s.includes('Completed')) return '<span class="badge b-done">✅ Done</span>';
  if(s.includes('Colors')) return '<span class="badge b-colors">🎨 Colors</span>';
  if(s.includes('Base')) return '<span class="badge b-base">🔨 Base</span>';
  return '<span class="badge b-pending">⏳ '+s+'</span>';
}
function websiteBadge(w){
  const no=!w||w==='No'||w===''||w==='false';
  return no?'<span class="badge b-nosite">No Site</span>':'<span class="badge b-yes">Has Site</span>';
}
function openBrief(id,biz){
  const p=SALES_PORTALS&&SALES_PORTALS[id];
  if(!p){alert('No portal data for this lead.');return;}
  function sec(obj,prefix){if(!obj)return'';const k=Object.keys(obj).find(k=>k.toUpperCase().includes(prefix.toUpperCase()));return k?obj[k]:'';}
  function m(text){if(!text)return'<p style="color:#2a4060;font-style:italic">&#x2014;</p>';if(typeof marked!=='undefined')return marked.parse(text);return'<div>'+text.replace(/&/g,'&amp;').replace(/\*\*(.*?)\*\*/g,'<strong>$1</strong>').replace(/\n/g,'<br>')+'</div>';}
  function esc(s){return(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');}
  // ── OVERVIEW data ──────────────────────────────────────────────────────────
  const snapRaw=sec(p.brief,'BUSINESS SNAPSHOT');
  const snapClean=snapRaw.split(/\n\*\*Review hooks/)[0].trim();
  const snapHtml=m(snapClean);
  const reviewMatches=[...snapRaw.matchAll(/\*"([^"]+)"\*/g)];
  const reviews=reviewMatches.slice(0,3).map(function(rm){return rm[1];});
  const compLandRaw=sec(p.brief,'COMPETITIVE LANDSCAPE');
  const compTableLines=compLandRaw.split('\n').filter(function(l){return l.startsWith('|')&&!l.includes('---')&&!/^\|\s*(Competitor|Comp\.?)\s*\|/i.test(l);});
  const compNames=compTableLines.slice(0,3).map(function(l){const cell=(l.split('|')[1]||'');return cell.replace(/\*\*/g,'').replace(/\s*\([^)]*\)/,'').trim();}).filter(Boolean);
  const step0Raw=sec(p.brief,'STEP 0')||sec(p.brief,'PRE-CALL DEMO');
  const frictionM=step0Raw.match(/\*\*Call friction angle:\*\*\s*(.*)/);
  const frictionAngle=frictionM?frictionM[1].trim():'';
  const frictionExplM=step0Raw.match(/\*\*Use this[^\n]*:\*\*\s*(.*)/);
  const frictionExpl=frictionExplM?frictionExplM[1].trim():'';
  const roiRaw=sec(p.brief,'COST OF A LOST');
  const roiLines=roiRaw.split('\n').filter(function(l){return l.startsWith('|')&&!l.includes('---')&&!/^\|\s*(Job Type)/i.test(l);});
  const bigJobRows=roiLines.filter(function(l){return!l.includes('missed jobs')&&!l.includes('3 missed');}).slice(0,4);
  const roiHtml=m(roiRaw);
  const mktHtml=m(sec(p.brief,'MARKET CONTEXT'));
  // Key observation from intel (for Battle Card pain point)
  const serpRawBC=p.has_intel?sec(p.intel,'KEYWORD DEMAND'):'';
  const keyObsBCM=serpRawBC?serpRawBC.match(/\*\*Key observation:\*\*([\s\S]*?)(?:\n\n|$)/):'';
  const keyObs=keyObsBCM?keyObsBCM[1].trim():'';
  // ── OUTREACH data ──────────────────────────────────────────────────────────
  const notesHtml=m(sec(p.brief,'NOTES'));
  const timingM=step0Raw.match(/^([\s\S]*?)(?=\n---)/);
  const sendTiming=timingM?timingM[1].trim():'';
  const msgM=step0Raw.match(/---\n([\s\S]+?)\n---/);
  const preCallMsg=msgM?msgM[1].trim():'';
  const bridgeM=step0Raw.match(/\*\*On the call[\s\S]*?(?=\*\*Opening —|$)/i);
  const callBridgeRaw=bridgeM?bridgeM[0].trim():'';
  const coldM=step0Raw.match(/\*\*Opening — cold[\s\S]*?(?=\n\*\*Call friction|\n> ⚠️|\n---\n|$)/i);
  const coldOpenRaw=coldM?coldM[0].trim():'';
  // ── STRATEGY data ──────────────────────────────────────────────────────────
  const callScriptRaw=sec(p.brief,'PITCH AMMO')||sec(p.brief,'CALL SCRIPT');
  const hookMatches=[...callScriptRaw.matchAll(/\*\*(Hook \d+[^*\n]+):\*\*[\s\S]*?(?:^|^> ?)(.+?)(?=\n\*\*Hook|\n\*\*The Close|$)/gm)];
  const hookRx=[...callScriptRaw.matchAll(/\*\*(Hook \d[^\*]+)\*\*[^\n]*\n>(.*)/g)];
  const pitchHooks=hookRx.slice(0,3).map(function(hm){return{label:hm[1].trim(),text:hm[2].trim()};});
  const anglesHtml=m(sec(PLAYBOOK,'PART 1'));
  // ── INTEL tab ──────────────────────────────────────────────────────────────
  let intelTabContent='';
  const intelTopPanels='<div class="two-col">'
    +'<div class="panel"><div class="panel-title">Market Context</div><div class="panel-body">'+mktHtml+'</div></div>'
    +'<div class="panel"><div class="panel-title">&#x1F4B5; Avg Job Values</div><div class="panel-body">'+roiHtml+'</div></div>'
    +'</div>';
  if(p.has_intel){
    const serpRaw=sec(p.intel,'KEYWORD DEMAND');
    const compRaw=sec(p.intel,"WHO'S WINNING");
    const auditRaw=sec(p.intel,'SITE AUDIT')||sec(p.intel,'CURRENT SITE');
    const bottomLineM=auditRaw.match(/\*\*Bottom line:\*\*([\s\S]*?)(?:\n\n|$)/);
    const bottomLine=bottomLineM?bottomLineM[1].trim():'';
    let auditTable='<table style="width:100%;border-collapse:collapse"><thead><tr>'
      +'<th style="text-align:left;padding:7px 10px;font-size:11px;letter-spacing:1px;color:#2a6080;border-bottom:1px solid rgba(0,212,255,.12);background:rgba(0,212,255,.04)">Signal</th>'
      +'<th style="text-align:left;padding:7px 10px;font-size:11px;letter-spacing:1px;color:#2a6080;border-bottom:1px solid rgba(0,212,255,.12);background:rgba(0,212,255,.04)">Status</th>'
      +'<th style="text-align:left;padding:7px 10px;font-size:11px;letter-spacing:1px;color:#2a6080;border-bottom:1px solid rgba(0,212,255,.12);background:rgba(0,212,255,.04)">Notes</th>'
      +'</tr></thead><tbody>';
    let auditRows=0;
    auditRaw.split('\n').forEach(function(l){
      if(!l.startsWith('|')||l.includes('---'))return;
      const cells=l.split('|').map(function(c){return c.trim();}).filter(function(c){return c;});
      if(cells.length<2||cells[0].toLowerCase()==='signal')return;
      const sig=cells[0]||'';const st=cells[1]||'';const notes=cells[2]||'';
      let stHtml='<span style="color:#6a90a8">'+esc(st)+'</span>';
      if(st.includes('❌'))stHtml='<span style="background:rgba(248,113,113,.12);color:#f87171;padding:2px 8px;border-radius:2px;font-weight:700;font-size:12px">❌ Missing</span>';
      else if(st.includes('✅'))stHtml='<span style="background:rgba(0,255,157,.1);color:#00ff9d;padding:2px 8px;border-radius:2px;font-weight:700;font-size:12px">✅ Present</span>';
      else if(st.includes('⚠'))stHtml='<span style="background:rgba(245,196,0,.1);color:#f5c400;padding:2px 8px;border-radius:2px;font-weight:700;font-size:12px">⚠️ Partial</span>';
      auditTable+='<tr><td style="padding:7px 10px;border-bottom:1px solid rgba(0,212,255,.05);font-size:13px;font-weight:600;color:#c8d8e8">'+esc(sig)+'</td>'
        +'<td style="padding:7px 10px;border-bottom:1px solid rgba(0,212,255,.05);white-space:nowrap">'+stHtml+'</td>'
        +'<td style="padding:7px 10px;border-bottom:1px solid rgba(0,212,255,.05);font-size:12px;color:#6a90a8">'+esc(notes)+'</td></tr>';
      auditRows++;
    });
    auditTable+='</tbody></table>';
    if(!auditRows)auditTable=m(auditRaw);
    const aiHtml=m(sec(p.intel,'AI VISIBILITY'));
    intelTabContent=intelTopPanels
      +'<div class="tier-full tier-banner">&#x1F7E2; Full Intel &#x2014; Tier 2. Walk through this on the video call.</div>'
      +(keyObs?'<div style="background:rgba(248,113,113,.08);border:1px solid rgba(248,113,113,.3);border-left:4px solid #f87171;border-radius:3px;padding:14px 16px;margin-bottom:16px">'
        +'<div style="font-size:11px;font-weight:600;letter-spacing:1.5px;color:#f87171;text-transform:uppercase;margin-bottom:6px">&#x26A0;&#xFE0F; Pain Point &#x2014; Key Observation</div>'
        +'<div style="font-size:14px;color:#e8c8c8;line-height:1.65">'+esc(keyObs)+'</div></div>':'')
      +'<div class="panel"><div class="panel-title">Current Site Audit</div><div class="panel-body">'+auditTable
      +(bottomLine?'<div style="background:rgba(248,113,113,.07);border:1px solid rgba(248,113,113,.25);border-left:3px solid #f87171;border-radius:2px;padding:10px 14px;margin-top:14px"><span style="font-size:11px;font-weight:700;letter-spacing:1px;color:#f87171;text-transform:uppercase">Bottom Line: </span><span style="font-size:13px;color:#c8a8a8">'+esc(bottomLine)+'</span></div>':'')
      +'</div></div>'
      +'<div class="panel"><div class="panel-title">Top Competitors</div><div class="panel-body">'+m(compRaw)+'</div></div>'
      +(aiHtml?'<div class="panel"><div class="panel-title">AI Visibility Assessment</div><div class="panel-body">'+aiHtml+'</div></div>':'');
  }else{
    intelTabContent=intelTopPanels
      +'<div class="tier-std tier-banner">&#x1F7E1; Standard Brief (Tier 1) &#x2014; Run /lead-intel after video call confirmed to unlock full intel.</div>'
      +'<div class="panel"><div class="panel-title">Competitive Landscape</div><div class="panel-body">'+m(compLandRaw)+'</div></div>';
  }
  // ── Badges ─────────────────────────────────────────────────────────────────
  const stageBadge=p.dev_stage&&p.dev_stage.includes('Completed')
    ?'<span style="background:rgba(0,255,157,.12);color:#00ff9d;border:1px solid rgba(0,255,157,.3);padding:2px 8px;border-radius:2px;font-size:11px;font-weight:600">&#x2705; Pitch Ready</span>'
    :p.dev_stage?'<span style="background:rgba(245,196,0,.1);color:#f5c400;border:1px solid rgba(245,196,0,.3);padding:2px 8px;border-radius:2px;font-size:11px">'+esc(p.dev_stage)+'</span>':'';
  const intelBadge=p.has_intel
    ?'<span style="background:rgba(0,255,157,.1);color:#00ff9d;border:1px solid rgba(0,255,157,.25);padding:2px 8px;border-radius:2px;font-size:11px;font-weight:600">&#x1F7E2; Full Intel</span>'
    :'<span style="background:rgba(245,196,0,.08);color:#f5c400;border:1px solid rgba(245,196,0,.2);padding:2px 8px;border-radius:2px;font-size:11px">&#x1F7E1; Standard Brief</span>';
  // ── CSS ────────────────────────────────────────────────────────────────────
  const css='*{box-sizing:border-box;margin:0;padding:0}:root{--c:#00d4ff;--g:#00ff9d;--y:#f5c400;--bg:#060b18;--panel:#0d1528;--border:rgba(0,212,255,.2)}body{font-family:\'Rajdhani\',sans-serif;background:var(--bg);color:#c8d8e8;min-height:100vh}body::before{content:\'\';position:fixed;inset:0;background-image:linear-gradient(rgba(0,212,255,.025) 1px,transparent 1px),linear-gradient(90deg,rgba(0,212,255,.025) 1px,transparent 1px);background-size:40px 40px;pointer-events:none;z-index:0}a{color:var(--c);text-decoration:none}.hdr{position:relative;z-index:10;background:linear-gradient(180deg,rgba(0,20,50,.98),rgba(6,11,24,.95));border-bottom:1px solid var(--border);padding:14px 24px;display:flex;justify-content:space-between;align-items:flex-start;gap:20px}.hdr-left{flex:1;min-width:0}.hdr-title{font-size:22px;font-weight:700;letter-spacing:2px;color:var(--c);text-shadow:0 0 12px rgba(0,212,255,.6);margin-bottom:6px}.hdr-meta{display:flex;gap:8px;align-items:center;flex-wrap:wrap;margin-bottom:8px}.hdr-contact{font-size:13px;color:#4a7090;display:flex;gap:16px;flex-wrap:wrap;margin-bottom:6px}.hdr-owner{font-size:12px;color:#3a6080;margin-bottom:2px}.hdr-actions{display:flex;flex-direction:column;gap:8px;align-items:flex-end;flex-shrink:0}.btn-demo{background:linear-gradient(90deg,rgba(0,255,157,.15),rgba(0,212,255,.1));border:1px solid rgba(0,255,157,.4);color:var(--g);padding:9px 18px;border-radius:3px;font-family:\'Rajdhani\',sans-serif;font-size:13px;font-weight:700;letter-spacing:1px;text-decoration:none;display:inline-block}.btn-site{background:rgba(245,196,0,.06);border:1px solid rgba(245,196,0,.3);color:#c8a840;padding:7px 14px;border-radius:3px;font-family:\'Rajdhani\',sans-serif;font-size:12px;font-weight:600;text-decoration:none;display:inline-block}.tab-nav{position:relative;z-index:10;display:flex;background:rgba(6,11,24,.95);border-bottom:1px solid var(--border);padding:0 20px;overflow-x:auto}.tab-btn{padding:11px 22px;font-size:13px;font-weight:600;letter-spacing:1px;text-transform:uppercase;color:#3a5570;border:none;background:transparent;cursor:pointer;border-bottom:2px solid transparent;transition:all .15s;white-space:nowrap;font-family:\'Rajdhani\',sans-serif}.tab-btn:hover{color:#0090cc}.tab-btn.active{color:var(--c);border-bottom-color:var(--c)}.tab-body{display:none;padding:20px 24px;position:relative;z-index:2;max-width:980px;margin:0 auto}.tab-body.active{display:block}#tab-strategy.active{display:flex;padding:0;max-width:none;margin:0;height:calc(100vh - 108px);overflow:hidden}.two-col{display:grid;grid-template-columns:1fr 1fr;gap:16px}@media(max-width:680px){.two-col{grid-template-columns:1fr}}.panel{background:var(--panel);border:1px solid var(--border);border-radius:4px;margin-bottom:16px;overflow:hidden;position:relative}.panel::before{content:\'\';position:absolute;top:0;left:0;right:0;height:1px;background:linear-gradient(90deg,transparent,var(--c),transparent);opacity:.3}.panel-title{font-size:11px;font-weight:600;letter-spacing:2px;text-transform:uppercase;color:#3a7090;padding:10px 16px 8px;border-bottom:1px solid rgba(0,212,255,.08)}.panel-body{padding:14px 16px}.panel-body h1,.panel-body h2,.panel-body h3{color:var(--c);margin:14px 0 6px;letter-spacing:.5px}.panel-body h1{font-size:17px}.panel-body h2{font-size:15px}.panel-body h3{font-size:14px;color:#8ab0c8}.panel-body p{margin-bottom:8px;font-size:13px;line-height:1.65}.panel-body ul{margin:6px 0 10px 22px;list-style:disc}.panel-body li{margin-bottom:4px;font-size:13px;line-height:1.5}.panel-body strong{color:var(--y)}.panel-body em{color:#a78bfa}.panel-body blockquote{border-left:3px solid rgba(0,212,255,.4);padding:6px 14px;margin:8px 0;color:#8ab0c8;font-size:13px;background:rgba(0,212,255,.03);border-radius:0 3px 3px 0}.panel-body table{width:100%;border-collapse:collapse;font-size:13px;margin:8px 0}.panel-body th{background:rgba(0,212,255,.06);padding:7px 10px;text-align:left;font-size:11px;font-weight:600;letter-spacing:1px;text-transform:uppercase;color:#2a6080;border-bottom:1px solid rgba(0,212,255,.1)}.panel-body td{padding:6px 10px;border-bottom:1px solid rgba(0,212,255,.05);font-size:13px}.panel-body tr:hover td{background:rgba(0,212,255,.03)}.panel-body hr{border:none;border-top:1px solid rgba(0,212,255,.1);margin:14px 0}.copy-block{background:#070e1c;border:1px solid rgba(0,255,157,.2);border-radius:3px;margin:12px 0}.copy-hdr{display:flex;justify-content:space-between;align-items:center;padding:8px 14px;border-bottom:1px solid rgba(0,255,157,.1)}.copy-hdr span{font-size:11px;letter-spacing:1px;text-transform:uppercase;color:#2a7060}.copy-btn{background:rgba(0,255,157,.1);border:1px solid rgba(0,255,157,.3);color:var(--g);padding:5px 14px;border-radius:2px;font-family:\'Rajdhani\',sans-serif;font-size:12px;font-weight:600;cursor:pointer;transition:all .15s}.copy-btn:hover{background:rgba(0,255,157,.2)}.copy-msg{padding:14px;font-size:13px;color:#8ab8a8;line-height:1.7;white-space:pre-wrap;font-family:\'Rajdhani\',sans-serif}.tier-banner{padding:10px 16px;margin-bottom:16px;border-radius:3px;font-size:13px;font-weight:600;letter-spacing:.5px}.tier-full{background:rgba(0,255,157,.07);border:1px solid rgba(0,255,157,.25);color:var(--g)}.tier-std{background:rgba(245,196,0,.07);border:1px solid rgba(245,196,0,.2);color:var(--y)}.otabs{display:flex;gap:6px;margin-bottom:14px;flex-wrap:wrap}.otab-btn{background:rgba(0,212,255,.05);border:1px solid rgba(0,212,255,.15);color:#3a6080;padding:7px 16px;border-radius:2px;font-family:\'Rajdhani\',sans-serif;font-size:12px;font-weight:600;cursor:pointer;letter-spacing:.5px;transition:all .15s}.otab-btn:hover{color:#0090cc;border-color:rgba(0,212,255,.3)}.otab-btn.active{background:rgba(0,212,255,.1);border-color:rgba(0,212,255,.5);color:var(--c)}.otab-pane{display:none}.otab-pane.active{display:block}.send-note{background:rgba(0,212,255,.04);border:1px solid rgba(0,212,255,.12);border-radius:3px;padding:10px 14px;margin-bottom:12px;font-size:12px;color:#4a8090;line-height:1.6}.comp-chip{background:rgba(0,212,255,.06);border:1px solid rgba(0,212,255,.2);color:#4a9090;padding:3px 10px;border-radius:2px;font-size:12px;font-weight:600;letter-spacing:.5px}.job-val-row{display:flex;justify-content:space-between;align-items:center;padding:8px 0;border-bottom:1px solid rgba(0,212,255,.06)}.job-val-row:last-child{border-bottom:none}.job-val-label{font-size:12px;color:#4a7080;max-width:55%}.job-val-amount{font-size:16px;font-weight:700;color:#00ff9d;text-align:right}.review-hook{font-size:12px;color:#7ab0b0;border-left:2px solid rgba(0,212,255,.25);padding:4px 10px;margin-bottom:6px;font-style:italic;line-height:1.5}.strat-wrap{display:flex;height:100%;overflow:hidden;width:100%}.strat-nav{width:220px;flex-shrink:0;background:#080d1a;border-right:1px solid rgba(0,212,255,.12);overflow-y:auto;display:flex;flex-direction:column}.strat-nav-inner{padding:10px 8px;flex:1}.strat-nav-label{font-size:9px;letter-spacing:2px;text-transform:uppercase;color:#2a4560;padding:10px 8px 4px}.snav-btn{display:block;width:100%;text-align:left;background:transparent;border:none;color:#3a6080;padding:7px 10px;border-radius:2px;font-family:\'Rajdhani\',sans-serif;font-size:12px;font-weight:600;cursor:pointer;letter-spacing:.5px;transition:all .12s;margin-bottom:1px}.snav-btn:hover{background:rgba(0,212,255,.06);color:#7ab0cc}.snav-btn.active{background:rgba(0,212,255,.1);color:var(--c);border-left:2px solid var(--c);padding-left:8px}.phase-dot{display:inline-block;width:8px;height:8px;border-radius:50%;margin-right:6px;vertical-align:middle}.strat-iframe-wrap{flex:1;overflow:hidden}.strat-iframe-wrap iframe{width:100%;height:100%;border:none}.pitch-ammo{padding:12px 10px;border-top:1px solid rgba(0,212,255,.1);margin-top:auto}.pitch-ammo-title{font-size:9px;letter-spacing:2px;text-transform:uppercase;color:#2a4560;margin-bottom:8px}.hook-card{background:rgba(0,20,50,.7);border:1px solid rgba(0,212,255,.1);border-radius:2px;padding:8px 10px;margin-bottom:6px}.hook-label{font-size:9px;letter-spacing:1px;text-transform:uppercase;color:#2a7060;margin-bottom:3px}.hook-text{font-size:11px;color:#7ab8a8;line-height:1.5;font-style:italic}.modal-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.7);z-index:100;align-items:center;justify-content:center}.modal-overlay.open{display:flex}.modal-box{background:#0d1528;border:1px solid var(--border);border-radius:6px;max-width:700px;width:90%;max-height:80vh;overflow-y:auto;padding:24px;position:relative}.modal-close{position:absolute;top:12px;right:16px;background:transparent;border:none;color:#3a6080;font-size:20px;cursor:pointer;font-family:\'Rajdhani\',sans-serif}.modal-title{font-size:14px;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:var(--c);margin-bottom:16px}.float-btn{position:fixed;bottom:24px;right:24px;background:linear-gradient(135deg,rgba(0,212,255,.2),rgba(192,132,252,.2));border:1px solid rgba(192,132,252,.4);color:#c084fc;padding:10px 16px;border-radius:24px;font-family:\'Rajdhani\',sans-serif;font-size:13px;font-weight:700;letter-spacing:1px;cursor:pointer;z-index:50;box-shadow:0 4px 20px rgba(192,132,252,.2)}';
  // ── Battle card & job value HTML ───────────────────────────────────────────
  const compChips=compNames.length?compNames.map(function(n){return'<span class="comp-chip">'+esc(n)+'</span>';}).join(' '):'<span style="color:#2a4060;font-size:12px">See Competitive Landscape tab</span>';
  const reviewsHtml=reviews.length?reviews.map(function(r){return'<div class="review-hook">&#x201C;'+esc(r)+'&#x201D;</div>';}).join(''):'';
  const jobValsHtml=bigJobRows.length?bigJobRows.map(function(row){
    const cells=row.split('|').map(function(c){return c.trim();}).filter(function(c){return c;});
    if(cells.length<2)return'';
    const lbl=cells[0].replace(/\*\*/g,'').trim();
    const val=cells[1].replace(/\*\*/g,'').trim();
    return'<div class="job-val-row"><span class="job-val-label">'+esc(lbl)+'</span><span class="job-val-amount">'+esc(val)+'</span></div>';
  }).join(''):'<div style="color:#2a4060;font-size:12px">See brief for job values</div>';
  // ── Strategy left nav ──────────────────────────────────────────────────────
  const phases=[
    {id:'overview',dot:'',label:'Pipeline Overview'},
    {id:'brief',dot:'',label:'Using the Brief'},
    {id:'p1',dot:'#a78bfa',label:'Phase 01: Open'},
    {id:'p2',dot:'#60a5fa',label:'Phase 02: Discovery'},
    {id:'p3',dot:'#f59e0b',label:'Phase 03: Deepen Pain'},
    {id:'p4',dot:'#00e5a0',label:'Phase 04: The Reveal'},
    {id:'p5',dot:'#f472b6',label:'Phase 05: Commitment'},
    {id:'p6',dot:'#ef4444',label:'Phase 06: Close / Bridge'},
    {id:'objections',dot:'',label:'Objections'},
    {id:'techniques',dot:'',label:'Techniques'}
  ];
  const navHtml=phases.map(function(ph,i){
    const dot=ph.dot?'<span class="phase-dot" style="background:'+ph.dot+'"></span>':'<span style="display:inline-block;width:14px"></span>';
    return'<button class="snav-btn'+(i===0?' active':'')+'" onclick="scriptNav(this,\''+ph.id+'\')">'+dot+esc(ph.label)+'</button>';
  }).join('');
  const pitchAmmoHtml=pitchHooks.length?pitchHooks.map(function(h){
    return'<div class="hook-card"><div class="hook-label">'+esc(h.label)+'</div><div class="hook-text">'+esc(h.text)+'</div></div>';
  }).join(''):'<div style="font-size:11px;color:#2a4060">No prospect hooks found — check the brief\'s Call Script section.</div>';
  // ── Build portal HTML ──────────────────────────────────────────────────────
  const portalHtml='<!DOCTYPE html><html><head><meta charset="UTF-8"/><title>'+esc(biz)+' &#x2014; Sales Portal</title>'
    +'<link href="https://fonts.googleapis.com/css2?family=Rajdhani:wght@400;500;600;700&display=swap" rel="stylesheet"/>'
    +'<style>'+css+'</style></head><body>'
    // Header
    +'<div class="hdr"><div class="hdr-left">'
    +'<div class="hdr-title">'+esc(biz)+'</div>'
    +'<div class="hdr-meta"><span style="background:rgba(0,212,255,.08);color:#4a9090;border:1px solid rgba(0,212,255,.2);padding:2px 9px;border-radius:2px;font-size:11px;letter-spacing:1px">'+esc(p.trade)+'</span>'
    +' <span style="font-size:13px;color:#f5c400">&#9733; '+esc(String(p.rating||''))+' <span style="color:#3a6080">('+esc(String(p.reviews||''))+' reviews)</span></span>'
    +' '+stageBadge+' '+intelBadge+'</div>'
    +'<div class="hdr-contact">'
    +(p.phone?'<span>&#x1F4DE; <a href="tel:'+p.phone.replace(/\D/g,'')+'" style="color:#00ff9d;font-weight:600">'+esc(p.phone)+'</a></span>':'')
    +(p.email&&p.email.toLowerCase()!='n/a'?'<span>&#x2709;&#xFE0F; <a href="mailto:'+p.email+'" style="color:#00d4ff">'+esc(p.email)+'</a></span>':'')
    +(p.address?'<span>&#x1F4CD; <a href="'+p.maps_url+'" target="_blank" style="color:#6aaa9a">'+esc(p.address)+'</a></span>':'')
    +(p.zip_income?'<span>&#x1F4B0; Area Income: <span style="color:#00ff9d;font-weight:600">'+esc(p.zip_income)+'</span></span>':'')
    +'</div>'
    +(p.owner?'<div class="hdr-owner">&#x1F464; Owner: <span style="color:#9ab8c8">'+esc(p.owner)+'</span></div>':'')
    +'</div><div class="hdr-actions">'
    +(p.demo_url?'<a href="'+p.demo_url+'" target="_blank" class="btn-demo">&#x1F310; Our Demo Site &#x2197;</a>':'')
    +((p.website_brief||p.website)?'<a href="'+(p.website_brief||p.website)+'" target="_blank" class="btn-site">&#x1F517; Current Prospect Website &#x2197;</a>':'')
    +'</div></div>'
    // Tab nav
    +'<div class="tab-nav">'
    +'<button class="tab-btn active" onclick="showTab(this,\'overview\')">&#x1F4CB; Overview</button>'
    +'<button class="tab-btn" onclick="showTab(this,\'outreach\')">&#x1F4E8; Outreach</button>'
    +'<button class="tab-btn" onclick="showTab(this,\'intel\')">&#x1F50D; Intel</button>'
    +'<button class="tab-btn" onclick="showTab(this,\'strategy\')">&#x26A1; Strategy</button>'
    +'</div>'
    // ── OVERVIEW TAB
    +'<div id="tab-overview" class="tab-body active">'
    +'<div class="two-col" style="margin-bottom:16px">'
    // Battle Card
    +'<div class="panel" style="border-color:rgba(192,132,252,.25)"><div class="panel-title" style="color:#c084fc">&#x26A1; Battle Card</div><div class="panel-body">'
    +'<div style="margin-bottom:12px"><div style="font-size:9px;letter-spacing:1.5px;text-transform:uppercase;color:#3a5070;margin-bottom:6px">Top Competitors</div>'
    +'<div style="display:flex;gap:6px;flex-wrap:wrap">'+compChips+'</div></div>'
    +(keyObs||frictionAngle?'<div style="margin-bottom:12px"><div style="font-size:9px;letter-spacing:1.5px;text-transform:uppercase;color:#3a5070;margin-bottom:4px">'+(keyObs?'&#x26A0;&#xFE0F; Key Intel Finding':'Their Biggest Gap')+'</div>'
      +'<div style="font-size:13px;color:#f5c400;font-weight:600;line-height:1.5">'+esc(keyObs||frictionAngle)+'</div>'
      +(!keyObs&&frictionExpl?'<div style="font-size:12px;color:#7a9090;margin-top:3px;line-height:1.5">'+esc(frictionExpl)+'</div>':'')
      +'</div>':'')
    +(reviewsHtml?'<div><div style="font-size:9px;letter-spacing:1.5px;text-transform:uppercase;color:#3a5070;margin-bottom:6px">Standout Reviews</div>'+reviewsHtml+'</div>':'')
    +'</div></div>'
    // Job Values
    +'<div class="panel" style="border-color:rgba(0,255,157,.2)"><div class="panel-title" style="color:#00ff9d">&#x1F4B5; What These Jobs Pay</div><div class="panel-body">'
    +jobValsHtml
    +'</div></div>'
    +'</div>'
    // Business Snapshot
    +'<div class="panel"><div class="panel-title">Business Snapshot</div><div class="panel-body">'+snapHtml+'</div></div>'
    +'</div>'
    // ── OUTREACH TAB
    +'<div id="tab-outreach" class="tab-body">'
    +'<div style="display:grid;grid-template-columns:35fr 65fr;gap:16px;align-items:start">'
    +'<div class="panel" style="margin-bottom:0"><div class="panel-title">Notes &amp; Best Call Times</div><div class="panel-body">'+notesHtml+'</div></div>'
    +'<div class="panel" style="margin-bottom:0"><div class="panel-title">Pre-Call Demo Drop</div><div class="panel-body">'
    +(sendTiming?'<div class="send-note">'+esc(sendTiming).replace(/\n/g,'<br>')+'</div>':'')
    +'<div class="otabs">'
    +(preCallMsg?'<button class="otab-btn active" onclick="showOTab(this,\'precall\')">&#x1F4AC; Pre-Call Message</button>':'')
    +(callBridgeRaw?'<button class="otab-btn'+(preCallMsg?'':' active')+'" onclick="showOTab(this,\'bridge\')">&#x260E;&#xFE0F; Call Bridge</button>':'')
    +(coldOpenRaw?'<button class="otab-btn'+(preCallMsg||callBridgeRaw?'':' active')+'" onclick="showOTab(this,\'cold\')">&#x2744;&#xFE0F; Cold Open</button>':'')
    +'</div>'
    +(preCallMsg?'<div class="otab-pane active" id="otab-precall"><div class="copy-block"><div class="copy-hdr"><span>Send via contact form / DM (24&#x2013;48h before call)</span><button class="copy-btn" id="cpbtn1" onclick="doCopy(\'cpbtn1\',\'ctxt1\')">&#x1F4CB; Copy</button></div><div class="copy-msg" id="ctxt1">'+esc(preCallMsg)+'</div></div></div>':'')
    +(callBridgeRaw?'<div class="otab-pane'+(preCallMsg?'':' active')+'" id="otab-bridge"><div class="copy-block"><div class="copy-hdr"><span>On-call bridge (after they saw the demo)</span><button class="copy-btn" id="cpbtn2" onclick="doCopy(\'cpbtn2\',\'ctxt2\')">&#x1F4CB; Copy</button></div><div class="copy-msg" id="ctxt2">'+esc(callBridgeRaw)+'</div></div></div>':'')
    +(coldOpenRaw?'<div class="otab-pane'+(preCallMsg||callBridgeRaw?'':' active')+'" id="otab-cold"><div class="copy-block"><div class="copy-hdr"><span>Cold open (no pre-call drop sent)</span><button class="copy-btn" id="cpbtn3" onclick="doCopy(\'cpbtn3\',\'ctxt3\')">&#x1F4CB; Copy</button></div><div class="copy-msg" id="ctxt3">'+esc(coldOpenRaw)+'</div></div></div>':'')
    +'</div></div>'
    +'</div>'
    +'</div>'
    // ── INTEL TAB
    +'<div id="tab-intel" class="tab-body">'+intelTabContent+'</div>'
    // ── STRATEGY TAB
    +'<div id="tab-strategy" class="tab-body">'
    +'<div class="strat-wrap">'
    +'<div class="strat-nav"><div class="strat-nav-inner">'
    +'<div class="strat-nav-label">Call Phases</div>'
    +navHtml
    +'</div>'
    +'<div class="pitch-ammo">'
    +'<div class="pitch-ammo-title">&#x1F3AF; Pitch Ammo</div>'
    +pitchAmmoHtml
    +'</div></div>'
    +'<div class="strat-iframe-wrap"><iframe id="stratIframe" src="./QF_Sales_Call_Script.html" style="width:100%;height:100%;border:none" onload="var d=this.contentDocument;if(d){var s=d.createElement(\'style\');s.textContent=\'.sidebar{display:none!important}.main{margin-left:0!important;padding-left:32px!important;max-width:none!important}\';d.head.appendChild(s);}"></iframe></div>'
    +'</div></div>'
    // ── Modal
    +'<div class="modal-overlay" id="whyModal"><div class="modal-box"><button class="modal-close" onclick="closeWhyModal()">&#x2715;</button><div class="modal-title">&#x1F4A1; Why a Website Matters in 2026</div><div class="panel-body">'+anglesHtml+'</div></div></div>'
    +'<button class="float-btn" onclick="openWhyModal()">&#x1F4A1; Why a Site?</button>'
    +'<script>function showTab(btn,id){document.querySelectorAll(".tab-body").forEach(function(t){t.style.display="";t.classList.remove("active")});document.querySelectorAll(".tab-btn").forEach(function(b){b.classList.remove("active")});document.getElementById("tab-"+id).classList.add("active");btn.classList.add("active");}function showOTab(btn,id){document.querySelectorAll(".otab-pane").forEach(function(p){p.classList.remove("active")});document.querySelectorAll(".otab-btn").forEach(function(b){b.classList.remove("active")});var el=document.getElementById("otab-"+id);if(el)el.classList.add("active");btn.classList.add("active");}function scriptNav(btn,id){var f=document.getElementById("stratIframe");if(f&&f.contentDocument){var el=f.contentDocument.getElementById(id);if(el)el.scrollIntoView({behavior:"smooth",block:"start"});}document.querySelectorAll(".snav-btn").forEach(function(b){b.classList.remove("active")});btn.classList.add("active");}function doCopy(btnId,txtId){var t=document.getElementById(txtId).innerText;if(navigator.clipboard){navigator.clipboard.writeText(t).then(function(){var b=document.getElementById(btnId);b.textContent="Copied!";setTimeout(function(){b.textContent="📋 Copy";},2000);});}else{var r=document.createRange();r.selectNodeContents(document.getElementById(txtId));window.getSelection().removeAllRanges();window.getSelection().addRange(r);document.execCommand("copy");}}function openWhyModal(){document.getElementById("whyModal").classList.add("open");}function closeWhyModal(){document.getElementById("whyModal").classList.remove("open");}document.getElementById("whyModal").addEventListener("click",function(e){if(e.target===this)closeWhyModal();});<\/script>'
    +'</body></html>';
  const w=window.open('','_blank');
  w.document.write(portalHtml);
  w.document.close();
}

function linkCell(url,label){
  if(!url||url==='—'||url==='') return '<span style="color:#2a4060">—</span>';
  return '<a href="'+url+'" target="_blank" style="color:#00d4ff;font-size:11px">'+label+' ↗</a>';
}
function phoneLink(p){
  if(!p) return '—';
  return '<a href="tel:'+p.replace(/\D/g,'')+'" style="color:#00ff9d;letter-spacing:.5px">'+p+'</a>';
}
function sortTable(id,col){
  const tbl=document.getElementById(id),tbody=tbl.querySelector('tbody');
  const rows=Array.from(tbody.querySelectorAll('tr'));
  const th=tbl.querySelectorAll('th')[col];
  const asc=th.dataset.sort!=='asc';
  tbl.querySelectorAll('th').forEach(h=>{h.dataset.sort=''});
  th.dataset.sort=asc?'asc':'desc';
  rows.sort((a,b)=>{
    const av=a.querySelectorAll('td')[col]?.innerText.trim()||'';
    const bv=b.querySelectorAll('td')[col]?.innerText.trim()||'';
    const an=parseFloat(av.replace(/[^0-9.-]/g,'')),bn=parseFloat(bv.replace(/[^0-9.-]/g,''));
    if(!isNaN(an)&&!isNaN(bn)) return asc?an-bn:bn-an;
    return asc?av.localeCompare(bv):bv.localeCompare(av);
  });
  rows.forEach(r=>tbody.appendChild(r));
}

// ── Chart defaults ────────────────────────────────────────────────────────────
Chart.defaults.color='#3a6080';
Chart.defaults.borderColor='rgba(0,212,255,.08)';
Chart.defaults.font.family='Rajdhani';

// ── OVERVIEW ─────────────────────────────────────────────────────────────────
function buildOverview(){
  document.getElementById('refreshed').textContent=REFRESHED;
  document.getElementById('kv-total').textContent=KPIS.total;
  document.getElementById('kv-yes').textContent=KPIS.yes;
  document.getElementById('kv-pitch').textContent=KPIS.pitch_ready;
  document.getElementById('kv-indev').textContent=KPIS.in_dev;
  document.getElementById('kv-pend').textContent=KPIS.pending_dev;
  document.getElementById('kv-maybe').textContent=KPIS.maybe;
  document.getElementById('kv-clients').textContent=KPIS.clients_won;

  // Funnel (no Prospected)
  new Chart(document.getElementById('ov-funnel'),{
    type:'bar',
    data:{
      labels:['Targeted','In Dev','Pitch Ready','Clients'],
      datasets:[{
        data:[KPIS.yes,KPIS.in_dev,KPIS.pitch_ready,KPIS.clients_won],
        backgroundColor:['rgba(192,132,252,.6)','rgba(245,196,0,.6)','rgba(0,255,157,.6)','rgba(167,139,250,.6)'],
        borderColor:['#c084fc','#f5c400','#00ff9d','#a78bfa'],
        borderWidth:1,borderRadius:3,borderSkipped:false
      }]
    },
    options:{plugins:{legend:{display:false}},scales:{
      y:{ticks:{color:'#2a5070',font:{size:10}},grid:{color:'rgba(0,212,255,.06)'}},
      x:{ticks:{color:'#3a6080',font:{size:9}},grid:{display:false}}
    }}
  });

  // Trade donut
  const tc=KPIS.trade_counts_yes;
  new Chart(document.getElementById('ov-trade'),{
    type:'doughnut',
    data:{
      labels:Object.keys(tc),
      datasets:[{
        data:Object.values(tc),
        backgroundColor:['rgba(0,212,255,.7)','rgba(0,255,157,.7)','rgba(245,196,0,.7)','rgba(255,107,53,.7)','rgba(167,139,250,.7)','rgba(0,144,204,.7)','rgba(79,70,229,.7)'],
        borderColor:'rgba(0,212,255,.15)',borderWidth:1
      }]
    },
    options:{plugins:{legend:{position:'right',labels:{color:'#4a7090',font:{size:10},boxWidth:10,padding:6}}},cutout:'60%'}
  });

  // Stage breakdown bars
  const stages=[
    {label:'Pitch Ready',val:KPIS.pitch_ready,max:KPIS.yes,color:'#00ff9d'},
    {label:'In Dev',val:KPIS.in_dev,max:KPIS.yes,color:'#f5c400'},
    {label:'Pending Dev',val:KPIS.pending_dev,max:KPIS.yes,color:'#4a6080'},
  ];
  const sb=document.getElementById('ov-stage-bars');
  stages.forEach(s=>{
    const pct=KPIS.yes>0?Math.round(s.val/KPIS.yes*100):0;
    sb.innerHTML+=`<div style="margin-bottom:10px">
      <div style="display:flex;justify-content:space-between;font-size:11px;letter-spacing:.5px;margin-bottom:4px">
        <span style="color:#4a7090">${s.label}</span>
        <span style="color:${s.color};font-family:'Share Tech Mono',monospace">${s.val} <span style="color:#2a4060">(${pct}%)</span></span>
      </div>
      <div style="height:4px;background:rgba(0,212,255,.08);border-radius:2px;overflow:hidden">
        <div style="width:${pct}%;height:100%;background:${s.color};border-radius:2px;box-shadow:0 0 6px ${s.color}88"></div>
      </div></div>`;
  });

  // Sales queue ranked list (left sidebar)
  const pitchLeads=LEADS.filter(l=>l['Dev Stage']==='✅ Completed');
  const sl=document.getElementById('ov-sales-list');
  pitchLeads.forEach((l,i)=>{
    const noWeb=!l['Has Website?']||l['Has Website?']==='No'||l['Has Website?']==='';
    sl.innerHTML+=`<div class="rank-item">
      <span class="rank-num">${String(i+1).padStart(2,'0')}</span>
      <div style="flex:1;min-width:0">
        <div style="font-size:12px;font-weight:600;color:#8ab0c8;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">${l['Business Name']||''}</div>
        <div style="font-size:10px;color:#2a5070">${l['Trade']||''} · ${l['ZIP / Area']||''}</div>
      </div>
      <div style="display:flex;gap:5px;flex-shrink:0;align-items:center">
        ${noWeb?'<span class="badge b-nosite">No Site</span>':''}
        ${l['Has Brief']?`<a href="#" onclick="openBrief('${l['Lead ID']}','${(l['Business Name']||'').replace(/'/g,"\\'")}');return false" style="color:#c084fc;font-size:10px" title="Sales Portal">🗂</a>`:''}
        <a href="${l['Maps URL']}" target="_blank" style="color:#00d4ff;font-size:10px" title="Google Maps">📍</a>
      </div>
    </div>`;
  });

  // Sales queue quick-access table (below map)
  const sqBody=document.getElementById('ov-sq-body');
  pitchLeads.forEach(l=>{
    const tr=document.createElement('tr');
    tr.innerHTML=`
      <td style="font-weight:600;color:#8ab0c8;white-space:nowrap">${l['Business Name']||''}</td>
      <td style="color:#4a7090">${l['Trade']||''}</td>
      <td>${l['Phone']?`<a href="tel:${l['Phone'].replace(/\D/g,'')}" style="color:#00ff9d">${l['Phone']}</a>`:'—'}</td>
      <td>${l['Staging URL']?`<a href="${l['Staging URL']}" target="_blank" style="color:#00d4ff;font-size:11px">🌐 Preview</a>`:'—'}</td>
      <td>${l['Has Brief']?`<a href="#" onclick="openBrief('${l['Lead ID']}','${(l['Business Name']||'').replace(/'/g,"\\'")}');return false" style="color:#c084fc;font-size:11px">🗂 Portal</a>`:'<span style="color:#2a4060">—</span>'}</td>
      <td><a href="${l['Maps URL']}" target="_blank" style="color:#4a9090;font-size:11px">📍 Map</a></td>`;
    sqBody.appendChild(tr);
  });

  // Insights
  const insights=[
    {icon:'🏆',color:'#f5c400',title:'Best Market: Charlotte NC',body:'96% qualified lead rate (48/50). Highest conversion of all 9 campaigns — strong opportunity to expand here.'},
    {icon:'💰',color:'#00ff9d',title:'$61,500 Backlog Potential',body:'41 yes-targeted leads are still pending dev. At $1,500/site that\'s $61,500 sitting in queue.'},
    {icon:'⚡',color:'#c084fc',title:'Electrical Gap',body:'Electrical is the #1 targeted trade (16 leads) but only 1 site is built. Biggest trade opportunity in the pipeline.'},
    {icon:'🌐',color:'#ff6b35',title:'8 Leads With No Website',body:'8 yes-targeted leads have zero web presence — easiest cold calls since you\'re not replacing anything.'},
    {icon:'📍',color:'#00d4ff',title:'NY-Heavy Pipeline',body:'54% of yes targets are in NY (31/57). CA, FL, and NC are underrepresented relative to lead quality found.'},
    {icon:'⭐',color:'#f5c400',title:'Avg Rating: 4.84★',body:'Your yes targets average 4.84 stars. You\'re pitching premium, well-reviewed businesses — use that credibility.'},
    {icon:'🏗️',color:'#f5c400',title:'50% Build Completion Rate',body:'8 of 16 started builds are finished. The other 8 are mid-dev — clearing them unlocks the next pitch cycle.'},
    {icon:'📊',color:'#00ff9d',title:'Floral Park: Best NY Market',body:'64% qualified lead rate in 11005 — top NY ZIP. Lower Manhattan follows at 54%.'},
    {icon:'🎯',color:'#c084fc',title:'Maybe Pool Has 32 Leads',body:'32 maybe leads averaging 4.83★. Plumbing (8) and Roofing (5) dominate — easy to promote to Yes with a quick review.'},
  ];
  const il=document.getElementById('ov-insights');
  insights.forEach(ins=>{
    il.innerHTML+=`<div style="padding:10px 12px;border-bottom:1px solid rgba(0,212,255,.06)">
      <div style="display:flex;align-items:center;gap:7px;margin-bottom:4px">
        <span style="font-size:14px">${ins.icon}</span>
        <span style="font-size:12px;font-weight:700;color:${ins.color};letter-spacing:.5px">${ins.title}</span>
      </div>
      <div style="font-size:11px;color:#4a7090;line-height:1.5;padding-left:21px">${ins.body}</div>
    </div>`;
  });

  // Campaign mini bar
  const cLabels=CAMPAIGN_LOG.map(r=>r['ZIP Code']);
  const cQual=CAMPAIGN_LOG.map(r=>parseInt(r['Qualified Leads'])||0);
  new Chart(document.getElementById('ov-campaign'),{
    type:'bar',data:{labels:cLabels,datasets:[{data:cQual,
      backgroundColor:'rgba(0,212,255,.45)',borderColor:'#00d4ff',borderWidth:1,borderRadius:3}]},
    options:{plugins:{legend:{display:false}},scales:{
      y:{ticks:{color:'#2a5070',font:{size:9}},grid:{color:'rgba(0,212,255,.05)'}},
      x:{ticks:{color:'#3a6080',font:{size:9}},grid:{display:false}}}}
  });

  // Overview map
  initOvMap();
}

// ── Overview Map ──────────────────────────────────────────────────────────────
let ovMap=null;
function initOvMap(){
  if(ovMap) return;
  ovMap=L.map('ov-map',{center:[37.5,-96],zoom:4,zoomControl:false,attributionControl:false});
  L.tileLayer('https://{s}.basemaps.cartocdn.com/dark_all/{z}/{x}/{y}{r}.png',{maxZoom:18}).addTo(ovMap);

  const colors={'pitch':'#00ff9d','indev':'#f5c400','pending':'#4f46e5','maybe':'#a78bfa','no':'#2a4060'};
  const heatPts=[];

  LEADS.forEach(l=>{
    if(!l.lat||!l.lng) return;
    const cat=getMarkerCat(l);
    if(cat==='no') return;
    const icon=L.divIcon({
      className:'',
      html:`<div style="width:10px;height:10px;border-radius:50%;background:${colors[cat]};border:1px solid rgba(255,255,255,.3);box-shadow:0 0 5px ${colors[cat]}88"></div>`,
      iconSize:[10,10],iconAnchor:[5,5]
    });
    const m=L.marker([l.lat,l.lng],{icon});
    m.bindPopup(popupHtml(l),{maxWidth:220});
    m.addTo(ovMap);
    heatPts.push([l.lat,l.lng,cat==='pitch'?1:cat==='indev'?.7:.4]);
  });

  L.heatLayer(heatPts,{radius:28,blur:18,gradient:{'0.2':'#3730a3','0.5':'#0090cc','0.8':'#00d4ff','1.0':'#00ff9d'}}).addTo(ovMap);
}

function getMarkerCat(l){
  if(l['Dev Stage']==='✅ Completed') return 'pitch';
  if(l['Dev Stage']==='🔨 Base Redesign to Template'||l['Dev Stage']==='🎨 Colors & Logos Updated') return 'indev';
  if(l['🎯 Target?']==='✅ Yes') return 'pending';
  if(l['🎯 Target?']==='🤔 Maybe') return 'maybe';
  return 'no';
}
function popupHtml(l){
  const noWeb=!l['Has Website?']||l['Has Website?']==='No'||l['Has Website?']==='';
  return `<div style="font-family:Rajdhani,sans-serif;min-width:180px">
    <div style="font-weight:700;font-size:13px;color:#c8d8e8;margin-bottom:3px">${l['Business Name']||''}</div>
    <div style="font-size:11px;color:#3a7090;margin-bottom:5px">${l['Trade']||''} · ${l['ZIP / Area']||''}</div>
    <div style="font-size:11px;margin-bottom:3px;color:#8ab0c8">★ ${l['Rating ★']||'?'} (${l['Reviews']||'?'} reviews)</div>
    <div style="font-size:11px;color:#4a9040;margin-bottom:3px">${l['Phone']||''}</div>
    ${noWeb?'<div style="color:#ff6b35;font-size:10px;margin-bottom:3px">⚠ No existing website</div>':''}
    ${l['Staging URL']?'<a href="'+l['Staging URL']+'" target="_blank" style="color:#00d4ff;font-size:11px">🌐 View Staging ↗</a>':''}
  </div>`;
}

// ── Dev Tracker ───────────────────────────────────────────────────────────────
let devLeads=[];
function buildDev(){
  devLeads=LEADS.filter(l=>l['🎯 Target?']==='✅ Yes');
  const trades=[...new Set(devLeads.map(l=>l['Trade']).filter(Boolean))].sort();
  const sel=document.getElementById('devTrade');
  trades.forEach(t=>{const o=document.createElement('option');o.value=t;o.textContent=t;sel.appendChild(o)});
  filterDev();
}
function filterDev(){
  const q=document.getElementById('devSearch').value.toLowerCase();
  const stage=document.getElementById('devStage').value;
  const trade=document.getElementById('devTrade').value;
  const noSite=isOn('devNoSite');
  let f=devLeads.filter(l=>{
    const hay=[l['Business Name'],l['Trade'],l['ZIP / Area'],l['Notes / Intel']].join(' ').toLowerCase();
    if(q&&!hay.includes(q)) return false;
    if(stage&&l['Dev Stage']!==stage) return false;
    if(trade&&l['Trade']!==trade) return false;
    if(noSite){const w=String(l['Has Website?']||'').toLowerCase();if(w==='yes'||w==='true') return false}
    return true;
  });
  document.getElementById('devCount').textContent=f.length+' LEADS';
  const tbody=document.querySelector('#devTable tbody');tbody.innerHTML='';
  f.forEach(l=>{
    const noWeb=!l['Has Website?']||l['Has Website?']==='No'||l['Has Website?']==='';
    const tr=document.createElement('tr');
    tr.innerHTML=`
      <td style="font-family:'Share Tech Mono',monospace;font-size:10px;color:#2a5070">${l['Lead ID']||''}</td>
      <td style="font-weight:600;color:#8ab0c8">${l['Business Name']||''}${noWeb?' <span class="badge b-nosite" style="margin-left:4px">No Site</span>':''}</td>
      <td style="color:#6a90a8">${l['Trade']||''}</td>
      <td style="font-size:11px;color:#3a6080">${l['ZIP / Area']||''}</td>
      <td>${phoneLink(l['Phone'])}</td>
      <td>${stageBadge(l['Dev Stage'])}</td>
      <td>${linkCell(l['Staging URL'],'View')}</td>
      <td style="font-size:11px;color:#3a5070">${l['Template']||'—'}</td>
      <td>${linkCell(l['Website'],'Site')}</td>`;
    tbody.appendChild(tr);
  });
}

// ── Sales Queue ───────────────────────────────────────────────────────────────
let salesLeads=[];
function buildSales(){
  salesLeads=LEADS.filter(l=>l['Dev Stage']==='✅ Completed');
  document.getElementById('revenuePill').textContent=
    '▶  '+salesLeads.length+' sites ready to pitch  ·  @ $5,000 avg  =  $'+(salesLeads.length*5000).toLocaleString()+' potential';
  const trades=[...new Set(salesLeads.map(l=>l['Trade']).filter(Boolean))].sort();
  const sel=document.getElementById('salesTrade');
  trades.forEach(t=>{const o=document.createElement('option');o.value=t;o.textContent=t;sel.appendChild(o)});
  filterSales();
}
function filterSales(){
  const q=document.getElementById('salesSearch').value.toLowerCase();
  const trade=document.getElementById('salesTrade').value;
  let f=salesLeads.filter(l=>{
    const hay=[l['Business Name'],l['Trade'],l['ZIP / Area']].join(' ').toLowerCase();
    if(q&&!hay.includes(q)) return false;
    if(trade&&l['Trade']!==trade) return false;
    return true;
  });
  document.getElementById('salesCount').textContent=f.length+' LEADS';
  const tbody=document.querySelector('#salesTable tbody');tbody.innerHTML='';
  f.forEach(l=>{
    const tr=document.createElement('tr');
    tr.innerHTML=`
      <td style="font-weight:600;color:#c8d8e8">${l['Business Name']||''}</td>
      <td style="color:#6a90a8">${l['Trade']||''}</td>
      <td style="font-size:11px;color:#3a6080">${l['ZIP / Area']||''}</td>
      <td>${phoneLink(l['Phone'])}</td>
      <td>${linkCell(l['Staging URL'],'🌐 Preview')}</td>
      <td>${l['Has Brief']?`<a href="#" onclick="openBrief('${l['Lead ID']}','${(l['Business Name']||'').replace(/'/g,"\\'")}');return false" style="color:#c084fc;font-size:12px;font-weight:600;display:block">🗂 Portal</a><span style="font-size:10px">${l['Has Intel']?'<span style="color:#00ff9d">🟢 Full Intel</span>':'<span style="color:#f5c400">🟡 Brief Only</span>'}</span>`:'<span style="color:#2a4060">—</span>'}</td>
      <td style="color:#f5c400;font-family:\'Share Tech Mono\',monospace">${l['Rating ★']||''}★</td>
      <td style="color:#6a90a8">${l['Reviews']||''}</td>
      <td><span class="badge b-base">To Call</span></td>`;
    tbody.appendChild(tr);
  });
}

// ── Active Pipeline ───────────────────────────────────────────────────────────
let pipeLeads=[];
function buildPipeline(){
  pipeLeads=LEADS.filter(l=>l['Dev Stage']==='🔨 Base Redesign to Template'||l['Dev Stage']==='🎨 Colors & Logos Updated');
  const trades=[...new Set(pipeLeads.map(l=>l['Trade']).filter(Boolean))].sort();
  const sel=document.getElementById('pipeTrade');
  trades.forEach(t=>{const o=document.createElement('option');o.value=t;o.textContent=t;sel.appendChild(o)});
  filterPipe();
}
function filterPipe(){
  const q=document.getElementById('pipeSearch').value.toLowerCase();
  const stage=document.getElementById('pipeStage').value;
  const trade=document.getElementById('pipeTrade').value;
  let f=pipeLeads.filter(l=>{
    const hay=[l['Business Name'],l['Trade'],l['ZIP / Area']].join(' ').toLowerCase();
    if(q&&!hay.includes(q)) return false;
    if(stage&&l['Dev Stage']!==stage) return false;
    if(trade&&l['Trade']!==trade) return false;
    return true;
  });
  document.getElementById('pipeCount').textContent=f.length+' ACTIVE BUILDS';
  const tbody=document.querySelector('#pipeTable tbody');tbody.innerHTML='';
  f.forEach(l=>{
    const tr=document.createElement('tr');
    tr.innerHTML=`
      <td style="font-weight:600;color:#c8d8e8">${l['Business Name']||''}</td>
      <td style="color:#6a90a8">${l['Trade']||''}</td>
      <td style="font-size:11px;color:#3a6080">${l['ZIP / Area']||''}</td>
      <td>${phoneLink(l['Phone'])}</td>
      <td>${stageBadge(l['Dev Stage'])}</td>
      <td style="font-size:11px;color:#3a5070">${l['Template']||'—'}</td>
      <td>${linkCell(l['Staging URL'],'Preview')}</td>
      <td>${linkCell(l['GitHub Repo'],'GitHub')}</td>
      <td>${linkCell(l['Website'],'Site')}</td>`;
    tbody.appendChild(tr);
  });
}

// ── Maybe ─────────────────────────────────────────────────────────────────────
let maybeLeads=[];
function buildMaybe(){
  maybeLeads=LEADS.filter(l=>l['🎯 Target?']==='🤔 Maybe')
    .sort((a,b)=>(parseFloat(b['Rating ★'])||0)-(parseFloat(a['Rating ★'])||0));
  const trades=[...new Set(maybeLeads.map(l=>l['Trade']).filter(Boolean))].sort();
  const sel=document.getElementById('maybeTrade');
  trades.forEach(t=>{const o=document.createElement('option');o.value=t;o.textContent=t;sel.appendChild(o)});
  filterMaybe();
}
function filterMaybe(){
  const q=document.getElementById('maybeSearch').value.toLowerCase();
  const trade=document.getElementById('maybeTrade').value;
  const noSite=isOn('maybeNoSite');
  let f=maybeLeads.filter(l=>{
    const hay=[l['Business Name'],l['Trade'],l['ZIP / Area'],l['Notes / Intel']].join(' ').toLowerCase();
    if(q&&!hay.includes(q)) return false;
    if(trade&&l['Trade']!==trade) return false;
    if(noSite){const w=String(l['Has Website?']||'').toLowerCase();if(w==='yes'||w==='true') return false}
    return true;
  });
  document.getElementById('maybeCount').textContent=f.length+' PROSPECTS';
  const tbody=document.querySelector('#maybeTable tbody');tbody.innerHTML='';
  f.forEach(l=>{
    const noWeb=!l['Has Website?']||l['Has Website?']==='No'||l['Has Website?']==='';
    const tr=document.createElement('tr');
    tr.innerHTML=`
      <td style="font-weight:600;color:#c8d8e8">${l['Business Name']||''}${noWeb?' <span class="badge b-nosite" style="margin-left:4px">No Site</span>':''}</td>
      <td style="color:#6a90a8">${l['Trade']||''}</td>
      <td style="font-size:11px;color:#3a6080">${l['ZIP / Area']||''}</td>
      <td style="color:#f5c400;font-family:\'Share Tech Mono\',monospace">${l['Rating ★']||''}★</td>
      <td style="color:#6a90a8">${l['Reviews']||''}</td>
      <td>${websiteBadge(l['Has Website?'])}</td>
      <td style="font-size:11px;color:#3a6080;max-width:220px">${l['Notes / Intel']||'—'}</td>
      <td>${linkCell(l['Google Maps'],'Map')}</td>`;
    tbody.appendChild(tr);
  });
}

// ── Campaign Charts ───────────────────────────────────────────────────────────
let campBuilt=false;
function buildCampaignCharts(){
  if(campBuilt) return; campBuilt=true;
  const tbody=document.querySelector('#campTable tbody'); tbody.innerHTML='';
  CAMPAIGN_LOG.forEach(r=>{
    const tr=document.createElement('tr');
    tr.innerHTML=`<td style="font-family:'Share Tech Mono',monospace;font-size:11px;color:#3a7090">${r['ZIP Code']||''}</td>
      <td style="color:#8ab0c8">${r['Location / Area']||''}</td>
      <td>${r['State']||''}</td>
      <td style="color:#00ff9d;font-weight:600">${r['Per Capita Income']||''}</td>
      <td style="font-size:11px;color:#3a6080">${r['Date Searched']||''}</td>
      <td style="font-size:11px;color:#3a5070;max-width:180px">${r['Trades Searched']||''}</td>
      <td style="color:#8ab0c8">${r['Leads Found']||''}</td>
      <td style="color:#f5c400;font-weight:600">${r['Qualified Leads']||''}</td>`;
    tbody.appendChild(tr);
  });
  const labels=CAMPAIGN_LOG.map(r=>r['ZIP Code']);
  const qualified=CAMPAIGN_LOG.map(r=>parseInt(r['Qualified Leads'])||0);
  const income=CAMPAIGN_LOG.map(r=>parseInt((r['Per Capita Income']||'0').replace(/[^0-9]/g,''))||0);
  new Chart(document.getElementById('campChart'),{
    type:'bar',data:{labels,datasets:[{label:'Qualified',data:qualified,
      backgroundColor:'rgba(0,212,255,.5)',borderColor:'#00d4ff',borderWidth:1,borderRadius:3}]},
    options:{plugins:{legend:{display:false}},scales:{
      y:{ticks:{color:'#2a5070'},grid:{color:'rgba(0,212,255,.06)'}},
      x:{ticks:{color:'#3a6080',font:{size:10}},grid:{display:false}}}}
  });
  new Chart(document.getElementById('incomeChart'),{
    type:'bar',data:{labels,datasets:[{label:'Income',data:income,
      backgroundColor:'rgba(0,255,157,.4)',borderColor:'#00ff9d',borderWidth:1,borderRadius:3}]},
    options:{plugins:{legend:{display:false}},scales:{
      y:{ticks:{color:'#2a5070',callback:v=>'$'+v.toLocaleString()},grid:{color:'rgba(0,212,255,.06)'}},
      x:{ticks:{color:'#3a6080',font:{size:10}},grid:{display:false}}}}
  });
}

// ── Full Map ──────────────────────────────────────────────────────────────────
let fullMap=null,fullHeatLayer=null,fullMarkersLayer=null,fullMapBuilt=false;
const mColors={'pitch':'#00ff9d','indev':'#f5c400','pending':'#4f46e5','maybe':'#a78bfa','no':'#2a4060'};

function initFullMap(){
  if(fullMapBuilt) return; fullMapBuilt=true;
  fullMap=L.map('full-map',{center:[37.5,-96],zoom:4});
  L.tileLayer('https://{s}.basemaps.cartocdn.com/dark_all/{z}/{x}/{y}{r}.png',{maxZoom:19,attribution:'© CARTO'}).addTo(fullMap);
  fullMarkersLayer=L.layerGroup().addTo(fullMap);
  // trade filter
  const trades=[...new Set(LEADS.map(l=>l['Trade']).filter(Boolean))].sort();
  const sel=document.getElementById('mapTrade');
  trades.forEach(t=>{const o=document.createElement('option');o.value=t;o.textContent=t;sel.appendChild(o)});
  renderFullMap();
}
function renderFullMap(){
  if(!fullMap) return;
  const tgt=document.getElementById('mapTarget').value;
  const trade=document.getElementById('mapTrade').value;
  const showNo=isOn('mapShowNo');
  fullMarkersLayer.clearLayers();
  const heatPts=[];
  LEADS.forEach(l=>{
    if(!l.lat||!l.lng) return;
    const cat=getMarkerCat(l);
    if(!showNo&&cat==='no') return;
    if(tgt==='yes'&&l['🎯 Target?']!=='✅ Yes') return;
    if(tgt==='maybe'&&l['🎯 Target?']!=='🤔 Maybe') return;
    if(tgt==='pitch'&&cat!=='pitch') return;
    if(trade&&l['Trade']!==trade) return;
    const col=mColors[cat];
    const icon=L.divIcon({
      className:'',
      html:`<div style="width:12px;height:12px;border-radius:50%;background:${col};border:2px solid rgba(255,255,255,.25);box-shadow:0 0 6px ${col}88"></div>`,
      iconSize:[12,12],iconAnchor:[6,6]
    });
    L.marker([l.lat,l.lng],{icon}).bindPopup(popupHtml(l),{maxWidth:240}).addTo(fullMarkersLayer);
    heatPts.push([l.lat,l.lng,cat==='pitch'?1:cat==='indev'?.7:cat==='maybe'?.5:.3]);
  });
  if(fullHeatLayer){fullMap.removeLayer(fullHeatLayer);fullHeatLayer=null}
  if(isOn('mapHeat')){
    fullHeatLayer=L.heatLayer(heatPts,{radius:30,blur:20,gradient:{'0.2':'#3730a3','0.5':'#0090cc','0.8':'#00d4ff','1.0':'#00ff9d'}}).addTo(fullMap);
  }
}
function toggleHeat(){
  if(!fullMap) return;
  if(fullHeatLayer){fullMap.removeLayer(fullHeatLayer);fullHeatLayer=null}
  if(isOn('mapHeat')) renderFullMap();
}

// ── Init all ──────────────────────────────────────────────────────────────────
buildOverview();
buildDev();
buildSales();
buildPipeline();
buildMaybe();
</script>
</body>
</html>
"""

html = HTML.replace('__DATA_PLACEHOLDER__', data_js)
with open(OUT_PATH,'w',encoding='utf-8') as f:
    f.write(html)
print(f"✅  {OUT_PATH}")
print(f"    Leads:{len(leads_clean)}  Dev:{len(dev_clean)}  Campaigns:{len(campaign_clean)}")
